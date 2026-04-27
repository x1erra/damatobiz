import streamlit as st
from io import BytesIO
from io import StringIO
import base64
import json
import os
import re
import warnings
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlparse
from uuid import uuid4
from zoneinfo import ZoneInfo
from zipfile import ZipFile

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go

try:
    from PIL import Image as PILImage
except ImportError:  # pragma: no cover
    PILImage = None


st.set_page_config(
    page_title="APP Look-Thru Reporting",
    page_icon="📊",
    layout="wide",
)


MANUAL_TEXT = """APP Look-Thru Reporting Manual

Purpose: reproduce the Portfolio Composition, Portfolio Breakdown, and Portfolio Diversification sections from holdings plus the monthly FactSet support files. The same row-level math is used throughout:

`Weighted MV = Parent Holding MV x Port. Weight / 100`

#### 1. Gather The Inputs

You need:

- Client holdings with `Fund Code`, `Fund Description`, `Total MV (CAD)`, and `Type`.
- Type must be `SAA`, `TAA`, or `SMA`.
- Monthly support files or ZIPs containing the `Composite Components` / `Port. Weight` tables.
- `Get Factset Model Codes` to identify the client model and support-file codes.
- `Asset Class Grouping For SMA` only when separately managed holdings are present.

Use the TAA holding to identify the model:

1. Find the client's TAA fund code in the holdings.
2. Look it up in `Get Factset Model Codes`.
3. Column A gives the FactSet model. Keep the full value, including the `C` or `P` suffix.
4. Column D gives the associated support / sales-charge code.
5. Use that support code to confirm which monthly support file is required.

Examples:

- `27017` uses support file `25017`
- `28016` uses support file `25016`

Do not assume the entered holdings code and support-file code are identical.

#### 2. Build The Holdings Sheet

Create one row per client holding:

| Fund Code | Fund Description | Total MV (CAD) | Type |
|-----------|------------------|----------------|------|
| 28000 | Fixed Income Managed Pool I | 262498.04 | SAA |
| 28016 | Tactical Asset Allocation Conservative Bal Pool I | 247645.78 | TAA |

Rules:

- Keep only rows with real market value.
- APP support-file rows are actively managed.
- SMA rows are separately managed.
- Portfolio Composition includes both actively managed and separately managed assets.
- Portfolio Breakdown and Portfolio Diversification exclude separately managed assets and use only actively managed market value as the denominator.

#### 3. Build The Detail Sheet

For each support file:

1. Open the file.
2. Find the table containing `Composite Components`, the report date, and `Port. Weight`.
3. Copy the block starting at `Non-Composite`.
4. Keep `Component` and `Port. Weight`.
5. Add the parent holding fields beside every copied row:
   - Parent Fund Code
   - Parent Fund Description
   - Parent MV
   - Type (`SAA` or `TAA`)
   - Support Code
   - Support File

The detail sheet should then have these working columns:

| Component | Port. Weight | Parent MV | Type | Composition Group | Breakdown Group | Diversification Group | Weighted MV | SAA MV | TAA MV |

Drag these formulas down every detail row:

```excel
Weighted MV = Parent MV * Port. Weight / 100
SAA MV      = IF(Type="SAA", Weighted MV, 0)
TAA MV      = IF(Type="TAA", Weighted MV, 0)
```

Do not double count parent rows and child rows. Use the report level described below for each section.

#### 4. Apply The Three Report Mappings

**A. Portfolio Composition mapping**

Use the high-level support rows plus fund-name classifications.

Classification:

- `Fixed Income` -> `Income`
- `Equity` -> `Equity`
- `Cash`, `Cash & Equivalents`, `[Cash]` -> `Cash`
- `Preferred`, `Derivatives`, `Currency Forwards`, `Commodities`, `FDS Outlier` -> `Other`
- private-market fund names -> `Private Alternatives`
- liquid alternative fund names -> `Private Alternatives`
- SMA rows come from `Asset Class Grouping For SMA`

Suggested Excel formula if using mapping tables:

```excel
=LET(
  x,TRIM([@Component]),
  base,XLOOKUP(x,AssetClassDetection[Label],AssetClassDetection[Composition Group],""),
  fund,XLOOKUP(x,FundNameDetection[Label],FundNameDetection[Composition Group],""),
  IF(fund<>"",fund,base)
)
```

Denominator: full portfolio market value.

```excel
Portfolio % = Group Weighted MV / Full Portfolio Market Value
```

**B. Portfolio Breakdown mapping**

Use actively managed support-file rows only. Exclude SMA / separately managed rows.

Classification:

- `Fixed Income` -> `Income`
- `Equity - International Equities` -> `International Equity`
- `Equity - US Equities` -> `US Equity`
- `Equity - Canadian Equities` -> `Canadian Equity`
- cash-like labels -> `Cash`
- preferred / derivatives / currency forwards / commodities / outlier labels -> `Other`
- private-market alternatives list -> `Alternatives`

Suggested Excel formula if using mapping tables:

```excel
=LET(
  x,TRIM([@Component]),
  XLOOKUP(x,BreakdownMapping[Label],BreakdownMapping[Breakdown Group],"")
)
```

Denominator: actively managed market value only.

```excel
Strategic % = SUMIFS(Weighted MV, Breakdown Group, row label, Type, "SAA") / Actively Managed MV
Tactical %  = SUMIFS(Weighted MV, Breakdown Group, row label, Type, "TAA") / Actively Managed MV
Portfolio % = Strategic % + Tactical %
```

**C. Portfolio Diversification mapping**

Use actively managed support-file rows only. Exclude SMA / separately managed rows.

This is sector-level look-through. It is not the same classification level as Portfolio Breakdown.

Priority order:

- contains `Government` -> `Government Bond`
- contains `Investment Grade` -> `Investment Grade Bond`
- contains `High Yield` -> `High Yield Bond`
- `Alternatives` and alternative/private fund names -> `Alternatives`
- `Cash`, `Cash & Equivalents`, `[Cash]` -> `Cash`
- equity sector labels at or above 2.00% stay as their sector:
  `Financials`, `Information Technology`, `Industrials`, `Consumer Discretionary`, `Health Care`, `Energy`, `Communication Services`, `Materials`
- everything else -> `Other`

Suggested Excel formula:

```excel
=LET(
  x,TRIM([@Component]),
  w,[@[Port. Weight]],
  sector,OR(x="Financials",x="Information Technology",x="Industrials",x="Consumer Discretionary",x="Health Care",x="Energy",x="Communication Services",x="Materials"),
  IF(ISNUMBER(SEARCH("Government",x)),"Government Bond",
  IF(ISNUMBER(SEARCH("Investment Grade",x)),"Investment Grade Bond",
  IF(ISNUMBER(SEARCH("High Yield",x)),"High Yield Bond",
  IF(ISNUMBER(XMATCH(x,AlternativeNames[Label])),"Alternatives",
  IF(OR(x="Alternatives",x="Cash",x="Cash & Equivalents",x="[Cash]"),IF(x="Alternatives","Alternatives","Cash"),
  IF(AND(sector,w>=2),x,"Other"))))))
)
```

Denominator: actively managed market value only.

```excel
Strategic % = SUMIFS(Weighted MV, Diversification Group, row label, Type, "SAA") / Actively Managed MV
Tactical %  = SUMIFS(Weighted MV, Diversification Group, row label, Type, "TAA") / Actively Managed MV
Portfolio % = Strategic % + Tactical %
```

#### 5. Denominators

Use the correct denominator or the numbers will look wrong even when the row math is correct.

| Section | Assets Included | Denominator |
|---------|-----------------|-------------|
| Portfolio Composition | Actively managed + separately managed | Full portfolio market value |
| Portfolio Breakdown | Actively managed only | Actively managed market value |
| Portfolio Diversification | Actively managed only | Actively managed market value |

Example:

- Full portfolio MV = `$734,598.79`
- Actively managed MV = `$686,064.78`
- Separately managed MV = `$48,534.01`

Private Alternatives on Portfolio Composition:

```excel
14287.17 / 734598.79 = 1.94%
```

Alternatives on Portfolio Breakdown:

```excel
14287.17 / 686064.78 = 2.08%
```

Alternatives on Portfolio Diversification:

```excel
44138.59 / 686064.78 = 6.43%
```

These are different because they come from different report levels and, for Composition, a different denominator.

#### 6. Build The Final Tables

**Portfolio Composition rows**

- Income
- Equity
- Balanced
- Cash
- Other
- Private Alternatives
- Total Market Value of Asset Classes

Show full portfolio MV, actively managed MV, separately managed MV, and `% of Portfolio`.

**Portfolio Breakdown rows**

- Income
- International Equity
- Canadian Equity
- US Equity
- Other
- Cash
- Alternatives
- Total of Actively Managed Assets

Sort rows by Portfolio % descending. Total stays last.

**Portfolio Diversification rows**

- Government Bond
- Investment Grade Bond
- Financials
- Information Technology
- Industrials
- Alternatives
- High Yield Bond
- Cash
- Other
- Materials
- Consumer Discretionary
- Energy
- Communication Services
- Health Care
- Total Portfolio Diversification

Sort rows by Portfolio % descending. Total stays last.

#### 7. Required Footer Notes

Use the support-file reporting period. Do not hardcode the month.

Portfolio Breakdown:

- `Asset Allocation sector weight is reported as of [support-file reporting month and year].`
- `The asset class 'Other' may include: Commodities, Derivatives, and/or Preferred Shares.`
- `The above portfolio breakdown does not include separately managed assets which are rebalanced separately from your actively managed assets.`

Portfolio Diversification:

- `Asset Allocation sector weight is reported as of [support-file reporting month and year].`
- `The category 'Other' may include: Commodities, Derivatives, Preferred Shares and/or any sector allocation under 2%.`
- `The above portfolio breakdown does not include separately managed assets which are rebalanced separately from your actively managed assets.`

#### 8. Validate Before Sending

Check these every time:

- Holdings total equals the Portfolio Composition total.
- Actively managed plus separately managed equals full portfolio MV.
- Portfolio Composition total equals 100.00%.
- Portfolio Breakdown total equals 100.00% of actively managed assets.
- Portfolio Diversification total equals 100.00% of actively managed assets.
- For every Breakdown and Diversification row: `Strategic % + Tactical % = Portfolio %`.
- `International Equity + US Equity + Canadian Equity` equals actively managed `Equity` from Portfolio Composition.
- All support files are from the same reporting period.
- Any variance over 3 basis points should be investigated.

#### 9. How To Audit A Challenged Number

Use the Excel export audit sheets:

- `Audit Summary` shows denominators and section totals.
- `Audit Composition` shows every row used for Portfolio Composition.
- `Audit Breakdown` shows every row used for Portfolio Breakdown.
- `Audit Diversification` shows every row used for Portfolio Diversification.

To explain a number:

1. Go to the relevant audit detail sheet.
2. Filter `Report Group` to the challenged row, such as `Alternatives`.
3. Review the source rows, support file, parent fund, type, port weight, and weighted MV.
4. Sum `SAA MV` and `TAA MV`.
5. Divide by the section denominator shown in `Audit Summary`.

This proves whether the difference is caused by:

- a different denominator,
- a different report level,
- SAA/TAA split,
- SMA exclusion,
- or a mapping/classification rule."""


MANUAL_FILE = Path(__file__).with_name("APP_Look_Thru_Calculation_Manual_Revised.md")
if MANUAL_FILE.exists():
    MANUAL_TEXT = MANUAL_FILE.read_text(encoding="utf-8")


COMPOSITION_GROUP_ORDER = [
    "Income",
    "Equity",
    "Balanced",
    "Liquid Alternatives",
    "Sector",
    "Cash",
    "Other",
    "Private Alternatives",
]

BREAKDOWN_GROUP_ORDER = [
    "Income",
    "International Equity",
    "US Equity",
    "Canadian Equity",
    "Cash",
    "Other",
    "Alternatives",
]

DIVERSIFICATION_GROUP_ORDER = [
    "Government Bond",
    "Investment Grade Bond",
    "Financials",
    "Information Technology",
    "High Yield Bond",
    "Alternatives",
    "Cash",
    "Industrials",
    "Consumer Discretionary",
    "Other",
    "Health Care",
    "Energy",
    "Communication Services",
    "Materials",
    "Consumer Staples",
    "Real Estate",
    "Utilities",
]

DIVERSIFICATION_SECTOR_GROUPS = {
    "Financials",
    "Information Technology",
    "Industrials",
    "Consumer Discretionary",
    "Health Care",
    "Energy",
    "Communication Services",
    "Materials",
    "Consumer Staples",
    "Real Estate",
    "Utilities",
}

ASSET_CLASS_DETECTION = {
    "EQUITY": "Equity",
    "CURRENCY FORWARDS": "Other",
    "FIXED INCOME": "Income",
    "CASH": "Cash",
    "COMMODITIES": "Other",
    "DERIVATIVES": "Other",
    "CASH & EQUIVALENTS": "Cash",
    "PREFERRED": "Other",
    "FDS OUTLIER": "Other",
    "[CASH]": "Cash",
    "ALTERNATIVES": "Private Alternatives",
    "CI LAWRENCE PARK ALTERNATIVE INVESTMENT GRADE CREDIT FUND": "Liquid Alt",
}

FUNDS_ALTERNATIVES_DETECTION = {
    "ALATE I L.P.": "Other",
    "ALATE I LP, RESTRICTED": "Private Alt",
    "AVENUE EUROPE SPECIAL SITUATIONS FUND V (U.S.), L.P.": "Private Alt",
    "AXIA U.S. GROCERY NET LEASE FUND I LP, RESTRICTED": "Private Alt",
    "CI ADAMS STREET GLOBAL PRIVATE MARKETS FUND (CLASS I)": "Private Alt",
    "CI ALTERNATIVE CREDIT OPPORTUNITIES FUND SERIES IN": "Other",
    "CI ALTERNATIVE INVESTMENT GRADE CREDIT FUND I": "Other",
    "CI ALTERNATIVE INVESTMENT GRADE CREDIT FUND (ETF C$ SERIES)": "Other",
    "CI LAWRENCE PARK ALTERNATIVE INVESTMENT GRADE CREDIT FUND": "Liquid Alt",
    "CI PM GROWTH FUND BL LP (SERIES I)": "Private Alt",
    "CI PRIVATE MARKETS GROWTH FUND I": "Private Alt",
    "CI PRIVATE MARKETS INCOME FUND (SERIES I)": "Other",
    "CI PRIVATE MARKETS GROWTH FUND - SERIES I INSTALLMENT RECEIPT": "Other",
    "CI PRIVATE MARKETS INCOME FUND - SERIES I INSTALLMENT RECEIPT": "Other",
    "HARBOURVEST ADELAIDE FEEDER E LP": "Private Alt",
    "HARBOURVEST ADELAIDE FEEDER F LP": "Private Alt",
    "HARBOURVEST ADELAIDE FEEDER G LP": "Private Alt",
    "HARBOURVEST INFRASTRUCTURE INCOME CAYMAN PARALLEL PARTNERSHIP L.": "Private Alt",
    "INSTITUTIONAL FIDUCIARY TR MONEY MKT PTF": "Equity",
    "INVESCO PREMIER US GOV MONEY PTF": "Other",
    "MONARCH CAPITAL PARTNERS OFFSHORE VI LP": "Private Alt",
    "MSILF PRIME PORTFOLIO-INST": "Equity",
    "T.RX CAPITAL FUND I, LP.": "Private Alt",
    "WHITEHORSE LIQUIDITY PARTNERS V LP": "Private Alt",
}

BREAKDOWN_ALTERNATIVES = {
    "ALTERNATIVES",
    "ALATE I LP, RESTRICTED",
    "AVENUE EUROPE SPECIAL SITUATIONS FUND V (U.S.), L.P.",
    "AXIA U.S. GROCERY NET LEASE FUND I LP, RESTRICTED",
    "CI ADAMS STREET GLOBAL PRIVATE MARKETS FUND (CLASS I)",
    "CI LAWRENCE PARK ALTERNATIVE INVESTMENT GRADE CREDIT FUND",
    "CI PM GROWTH FUND BL LP (SERIES I)",
    "CI PRIVATE MARKETS GROWTH FUND I",
    "DEMOPOLIS EQUITY PARTNERS FUND I, L.P.",
    "HARBOURVEST ADELAIDE FEEDER E LP",
    "HARBOURVEST ADELAIDE FEEDER F LP",
    "HARBOURVEST ADELAIDE FEEDER G LP",
    "HARBOURVEST INFRASTRUCTURE INCOME CAYMAN PARALLEL PARTNERSHIP L.",
    "MONARCH CAPITAL PARTNERS OFFSHORE VI LP",
    "T.RX CAPITAL FUND I, LP.",
    "WHITEHORSE LIQUIDITY PARTNERS V LP",
}

BREAKDOWN_CASH = {
    "CASH & EQUIVALENTS",
    "INSTITUTIONAL FIDUCIARY TR MONEY MKT PTF",
    "MSILF PRIME PORTFOLIO-INST",
    "[CASH]",
}

BREAKDOWN_OTHER = {
    "ALATE I L.P.",
    "CI ALTERNATIVE CREDIT OPPORTUNITIES FUND SERIES IN",
    "COMMODITIES",
    "PREFERRED",
    "CURRENCY FORWARDS",
    "DERIVATIVES",
    "FDS OUTLIER",
    "CI ALTERNATIVE INVESTMENT GRADE CREDIT FUND I",
    "CI ALTERNATIVE INVESTMENT GRADE CREDIT FUND (ETF C$ SERIES)",
    "CI PRIVATE MARKETS INCOME FUND (SERIES I)",
    "CI PRIVATE MARKETS GROWTH FUND - SERIES I INSTALLMENT RECEIPT",
    "CI PRIVATE MARKETS INCOME FUND - SERIES I INSTALLMENT RECEIPT",
    "INVESCO PREMIER US GOV MONEY PTF",
}

BREAKDOWN_DIRECT = {
    "EQUITY - CANADIAN EQUITIES": "Canadian Equity",
    "EQUITY - INTERNATIONAL EQUITIES": "International Equity",
    "EQUITY - US EQUITIES": "US Equity",
    "FIXED INCOME": "Income",
}

DIVERSIFICATION_DIRECT = {
    "GOVERNMENT": "Government Bond",
    "INVESTMENT GRADE": "Investment Grade Bond",
    "HIGH YIELD": "High Yield Bond",
    "FINANCIALS": "Financials",
    "INFORMATION TECHNOLOGY": "Information Technology",
    "INDUSTRIALS": "Industrials",
    "CONSUMER DISCRETIONARY": "Consumer Discretionary",
    "HEALTH CARE": "Health Care",
    "ENERGY": "Energy",
    "COMMUNICATION SERVICES": "Communication Services",
    "MATERIALS": "Materials",
    "ALTERNATIVES": "Alternatives",
    "CASH": "Cash",
    "CASH & EQUIVALENTS": "Cash",
    "[CASH]": "Cash",
    "COMMODITIES": "Other",
    "PREFERRED": "Other",
    "DERIVATIVES": "Other",
    "CURRENCY FORWARDS": "Other",
    "FDS OUTLIER": "Other",
    "CONSUMER STAPLES": "Consumer Staples",
    "REAL ESTATE": "Real Estate",
    "UTILITIES": "Utilities",
    "OTHER": "Other",
}

DIVERSIFICATION_WRAPPERS = {
    "NON-COMPOSITE",
    "FUNDS",
    "DIRECT",
    "EQUITY",
    "FIXED INCOME",
    "CORPORATE",
    "EQUITY - CANADIAN EQUITIES",
    "EQUITY - INTERNATIONAL EQUITIES",
    "EQUITY - US EQUITIES",
}

PRIMARY_BLUE = "#25356B"
SECONDARY_GRAY = "#9C9CA6"
BASE_DIR = Path(__file__).resolve().parent
REFERENCE_DIR = BASE_DIR / "Reference"
DATA_DIR = Path(os.environ.get("APP_DATA_DIR", BASE_DIR))
DATA_DIR.mkdir(parents=True, exist_ok=True)
APP_TIMEZONE = ZoneInfo(os.environ.get("APP_TIMEZONE", "America/Toronto"))
MANUAL_HOLDINGS_COLUMNS = ["Fund Code", "Fund Description", "Total MV (CAD)", "saa_taa"]
HOLDING_TYPE_OPTIONS = ["SAA", "TAA", "SMA"]
DEFAULT_HOLDINGS_INPUT = pd.DataFrame(
    [
        {
            "Fund Code": "",
            "Fund Description": "",
            "Total MV (CAD)": None,
            "saa_taa": "SAA",
        }
    ]
)
DRAFT_PATH = DATA_DIR / ".portfolio_composition_draft.json"
HISTORY_PATH = DATA_DIR / ".portfolio_account_history.json"
HISTORY_LIMIT = 50
FACTSET_MODEL_COLUMNS = ["factset_model_code", "sales_charge_code", "mandate_code", "fund_legal_name", "saa_taa"]
SMA_GROUPING_PATH = REFERENCE_DIR / "Asset Class Grouping For SMA.csv"
SMA_GROUPING_NUMBERS_PATH = REFERENCE_DIR / "Asset Class Grouping For SMA.numbers"
HOLDINGS_TEXT_TEMPLATE = ""


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_key(value: object) -> str:
    return normalize_text(value).upper()


def normalize_header(value: object) -> str:
    return re.sub(r"[^A-Z0-9]+", "", normalize_key(value))


def normalize_code(value: object) -> Optional[str]:
    text = normalize_text(value)
    if not text:
        return None
    match = re.search(r"(\d+)", text)
    if not match:
        return None
    return str(int(match.group(1)))


def coerce_number_series(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace("$", "", regex=False)
        .str.replace("%", "", regex=False)
        .str.strip()
    )
    cleaned = cleaned.replace({"": pd.NA, "nan": pd.NA, "None": pd.NA, "-": pd.NA})
    return pd.to_numeric(cleaned, errors="coerce")


def extract_mandate_code(value: object) -> Optional[str]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = normalize_text(value)
    match = re.search(r"(\d{5})", text)
    return match.group(1) if match else None


def normalize_holding_type(value: object) -> str:
    holding_type = normalize_key(value)
    if holding_type in HOLDING_TYPE_OPTIONS:
        return holding_type
    return ""


def allocation_bucket_for_holding_type(value: object) -> str:
    holding_type = normalize_holding_type(value)
    if holding_type == "TAA":
        return "TAA"
    if holding_type in {"SAA", "SMA"}:
        return "SAA"
    return ""


def infer_model_series_suffix(description: object) -> str:
    text = f" {normalize_key(description)} "
    if any(token in text for token in [" POOL ", " POOL I ", " POOL F "]):
        return "P"
    if any(token in text for token in [" CLASS ", " CLASS F ", " CC ", " CORPORATE "]):
        return "C"
    return ""


def get_latest_reference_file(stem: str, suffixes: Iterable[str]) -> Optional[Path]:
    candidates: List[Path] = []
    stem_key = normalize_key(stem)
    suffix_set = {suffix.lower() for suffix in suffixes}
    if not REFERENCE_DIR.exists():
        return None
    for path in REFERENCE_DIR.iterdir():
        if not path.is_file():
            continue
        if path.suffix.lower() not in suffix_set:
            continue
        base_name = path.stem
        while Path(base_name).suffix.lower() in suffix_set:
            base_name = Path(base_name).stem
        base_name = re.sub(r"\s+\(\d+\)$", "", base_name)
        if normalize_key(base_name) == stem_key:
            candidates.append(path)
    if not candidates:
        return None
    return max(candidates, key=lambda path: (path.stat().st_mtime, path.name))


@st.cache_data(show_spinner=False)
def load_numbers_table(filepath: str) -> pd.DataFrame:
    try:
        from numbers_parser import Document
    except ImportError as exc:
        raise ImportError(
            "Reading bundled Numbers reference files requires the `numbers-parser` package."
        ) from exc

    path = Path(filepath)
    if not path.exists():
        return pd.DataFrame()

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=RuntimeWarning, module="numbers_parser")
        document = Document(path)

    if not document.sheets or not document.sheets[0].tables:
        return pd.DataFrame()

    rows = document.sheets[0].tables[0].rows(values_only=True)
    if not rows:
        return pd.DataFrame()

    header = [normalize_text(value) for value in rows[0]]
    records = [row for row in rows[1:] if any(value is not None and normalize_text(value) != "" for value in row)]
    if not records:
        return pd.DataFrame(columns=header)

    return pd.DataFrame(records, columns=header)


@st.cache_data(show_spinner=False)
def load_csv_table(filepath: str) -> pd.DataFrame:
    path = Path(filepath)
    if not path.exists():
        return pd.DataFrame()
    return pd.read_csv(path)


@st.cache_data(show_spinner=False)
def load_reference_table(filepath: str) -> pd.DataFrame:
    suffix = Path(filepath).suffix.lower()
    if suffix == ".csv":
        return load_csv_table(filepath)
    if suffix == ".numbers":
        return load_numbers_table(filepath)
    if suffix in {".xlsx", ".xls", ".xlsm"}:
        path = Path(filepath)
        return read_excel_sheet(path.read_bytes(), path.name, sheet_name=0, header=0)
    return pd.DataFrame()


@st.cache_data(show_spinner=False)
def load_numbers_table_from_bytes(file_bytes: bytes, filename: str) -> pd.DataFrame:
    try:
        from numbers_parser import Document
    except ImportError as exc:
        raise ImportError(
            "Reading uploaded Numbers reference files requires the `numbers-parser` package."
        ) from exc

    import tempfile

    suffix = Path(filename).suffix or ".numbers"
    with tempfile.NamedTemporaryFile(delete=True, suffix=suffix) as tmp:
        tmp.write(file_bytes)
        tmp.flush()
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=RuntimeWarning, module="numbers_parser")
            document = Document(tmp.name)

    if not document.sheets or not document.sheets[0].tables:
        return pd.DataFrame()

    rows = document.sheets[0].tables[0].rows(values_only=True)
    if not rows:
        return pd.DataFrame()

    header = [normalize_text(value) for value in rows[0]]
    records = [row for row in rows[1:] if any(value is not None and normalize_text(value) != "" for value in row)]
    if not records:
        return pd.DataFrame(columns=header)

    return pd.DataFrame(records, columns=header)


@st.cache_data(show_spinner=False)
def load_reference_table_from_bytes(file_bytes: bytes, filename: str) -> pd.DataFrame:
    suffix = Path(filename).suffix.lower()
    if suffix == ".numbers":
        return load_numbers_table_from_bytes(file_bytes, filename)
    if suffix == ".csv":
        return pd.read_csv(BytesIO(file_bytes))
    if suffix in {".xlsx", ".xls", ".xlsm"}:
        return read_tabular_sheet(file_bytes, filename, sheet_name=0, header=0)
    raise ValueError(f"Unsupported reference file type for `{filename}`.")


def normalize_factset_model_codes_table(table: pd.DataFrame) -> pd.DataFrame:
    if table.empty:
        return pd.DataFrame(columns=FACTSET_MODEL_COLUMNS)

    normalized = {col: normalize_header(col) for col in table.columns}
    rename_map = {}
    for col, key in normalized.items():
        if key == "FACTSETMODELCODE":
            rename_map[col] = "factset_model_code"
        elif key == "SALESCHARGECODE":
            rename_map[col] = "sales_charge_code"
        elif key == "MANDATECODE":
            rename_map[col] = "mandate_code"
        elif key == "FUNDLEGALNAME":
            rename_map[col] = "fund_legal_name"
        elif key == "SAATAA":
            rename_map[col] = "saa_taa"
    table = table.rename(columns=rename_map)

    required_cols = {"factset_model_code", "sales_charge_code", "mandate_code"}
    if not required_cols.issubset(set(table.columns)):
        return pd.DataFrame(columns=FACTSET_MODEL_COLUMNS)

    keep_cols = FACTSET_MODEL_COLUMNS
    table = table[[col for col in keep_cols if col in table.columns]].copy()
    for column in keep_cols:
        if column not in table.columns:
            table[column] = ""

    if "sales_charge_code" in table.columns:
        table["sales_charge_code"] = table["sales_charge_code"].apply(normalize_code)
    if "mandate_code" in table.columns:
        table["mandate_code"] = table["mandate_code"].apply(normalize_code)
    if "factset_model_code" in table.columns:
        table["factset_model_code"] = table["factset_model_code"].apply(normalize_text)
    if "fund_legal_name" in table.columns:
        table["fund_legal_name"] = table["fund_legal_name"].apply(normalize_text)
    if "saa_taa" in table.columns:
        table["saa_taa"] = table["saa_taa"].apply(normalize_text).str.upper()
    table = table.dropna(subset=["factset_model_code", "sales_charge_code"]).copy()
    return table.reset_index(drop=True)


@st.cache_data(show_spinner=False)
def load_factset_model_codes(override_bytes: Optional[bytes] = None, override_filename: str = "") -> pd.DataFrame:
    if override_bytes:
        table = load_reference_table_from_bytes(override_bytes, override_filename or "factset_model_codes.xlsx")
        return normalize_factset_model_codes_table(table)

    source_path = get_latest_reference_file("Get Factset Model Codes", [".csv", ".xlsx", ".xls", ".xlsm"])
    if source_path is None:
        source_path = get_latest_reference_file("Get Factset Model Codes", [".numbers"])
    if source_path is None:
        return pd.DataFrame(columns=FACTSET_MODEL_COLUMNS)
    table = load_reference_table(str(source_path))
    return normalize_factset_model_codes_table(table)


@st.cache_data(show_spinner=False)
def load_sma_grouping_table(override_bytes: Optional[bytes] = None, override_filename: str = "") -> pd.DataFrame:
    if override_bytes:
        table = load_reference_table_from_bytes(override_bytes, override_filename or "sma_override.xlsx")
    else:
        source_path = get_latest_reference_file(
            "Asset Class Grouping For SMA",
            [".csv", ".xlsx", ".xls", ".xlsm"],
        )
        if source_path is None:
            source_path = get_latest_reference_file("Asset Class Grouping For SMA", [".numbers"])
        if source_path is None:
            source_path = SMA_GROUPING_PATH if SMA_GROUPING_PATH.exists() else SMA_GROUPING_NUMBERS_PATH
        table = load_reference_table(str(source_path))
    if table.empty:
        return pd.DataFrame(
            columns=[
                "Fund Code",
                "Sales Charge Code.Legal Name",
                "Portfolio Composition",
                "Portfolio Breakdown",
                "Portfolio AAbA",
            ]
        )

    normalized = {col: normalize_header(col) for col in table.columns}
    rename_map = {}
    for col, key in normalized.items():
        if key == "FUNDCODE":
            rename_map[col] = "Fund Code"
        elif key == "SALESCHARGECODELEGALNAME":
            rename_map[col] = "Sales Charge Code.Legal Name"
        elif key == "PORTFOLIOCOMPOSITION":
            rename_map[col] = "Portfolio Composition"
        elif key == "PORTFOLIOBREAKDOWN":
            rename_map[col] = "Portfolio Breakdown"
        elif key == "PORTFOLIOAABA":
            rename_map[col] = "Portfolio AAbA"
    table = table.rename(columns=rename_map)
    keep_cols = [
        "Fund Code",
        "Sales Charge Code.Legal Name",
        "Portfolio Composition",
        "Portfolio Breakdown",
        "Portfolio AAbA",
    ]
    table = table[[col for col in keep_cols if col in table.columns]].copy()
    table["Fund Code"] = table["Fund Code"].apply(normalize_code)
    for col in ["Sales Charge Code.Legal Name", "Portfolio Composition", "Portfolio Breakdown", "Portfolio AAbA"]:
        if col in table.columns:
            table[col] = table[col].apply(normalize_text)
    table = table[table["Fund Code"].notna()].drop_duplicates(subset=["Fund Code"], keep="first").reset_index(drop=True)
    return table


def get_sma_grouping_table(sma_override: Optional[dict] = None) -> pd.DataFrame:
    if sma_override:
        return load_sma_grouping_table(
            override_bytes=sma_override["bytes"],
            override_filename=sma_override["filename"],
        )
    return load_sma_grouping_table()


def get_factset_model_table(factset_override: Optional[dict] = None) -> pd.DataFrame:
    if factset_override:
        return load_factset_model_codes(
            override_bytes=factset_override["bytes"],
            override_filename=factset_override["filename"],
        )
    return load_factset_model_codes()


def validate_factset_model_file(record: dict) -> Tuple[bool, str]:
    try:
        table = get_factset_model_table(record)
    except Exception as exc:
        return False, f"FactSet model codes file could not be read: {exc}"
    if table.empty:
        return (
            False,
            "FactSet model codes file was ignored because it does not contain the required columns: FactSet Model Code, Sales Charge Code, and Mandate Code.",
        )
    return True, f"FactSet model codes loaded: {record['filename']} ({len(table):,} rows)."


def validate_sma_grouping_file(record: dict) -> Tuple[bool, str]:
    try:
        table = get_sma_grouping_table(record)
    except Exception as exc:
        return False, f"SMA grouping file could not be read: {exc}"
    required_cols = {"Fund Code", "Portfolio Composition", "Portfolio Breakdown"}
    if table.empty or not required_cols.issubset(set(table.columns)):
        return (
            False,
            "SMA grouping file was ignored because it does not contain the required columns: Fund Code, Portfolio Composition, and Portfolio Breakdown.",
        )
    return True, f"SMA grouping loaded: {record['filename']} ({len(table):,} rows)."


def normalize_uploaded_record(uploaded: Optional[object]) -> Optional[dict]:
    if uploaded is None:
        return None
    return {
        "filename": uploaded.name,
        "bytes": uploaded.getvalue(),
    }


def decode_saved_record(record: Optional[dict]) -> Optional[dict]:
    if not record:
        return None
    filename = record.get("filename")
    content_b64 = record.get("content_b64")
    if not filename or not content_b64:
        return None
    try:
        return {
            "filename": filename,
            "bytes": base64.b64decode(content_b64.encode("ascii")),
        }
    except Exception:
        return None


def encode_saved_record(record: Optional[dict]) -> Optional[dict]:
    if not record:
        return None
    return {
        "filename": record["filename"],
        "content_b64": base64.b64encode(record["bytes"]).decode("ascii"),
    }


def get_excel_engine(filename: str) -> Optional[str]:
    lower = filename.lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return "openpyxl"
    if lower.endswith(".xls"):
        return "xlrd"
    return None


def is_csv_file(filename: str) -> bool:
    return filename.lower().endswith(".csv")


@st.cache_data(show_spinner=False)
def read_excel_sheet(file_bytes: bytes, filename: str, sheet_name: object, header=None) -> pd.DataFrame:
    engine = get_excel_engine(filename)
    if engine is None:
        raise ValueError(f"Unsupported Excel file type for `{filename}`.")
    try:
        return pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, header=header, engine=engine)
    except ImportError as exc:
        if engine == "openpyxl":
            raise ImportError(
                "Reading .xlsx files requires the `openpyxl` package. Install it with `pip install openpyxl`."
            ) from exc
        if engine == "xlrd":
            raise ImportError(
                "Reading .xls files requires the optional `xlrd` package. Convert the file to .xlsx or install xlrd."
            ) from exc
        raise


@st.cache_data(show_spinner=False)
def read_tabular_sheet(file_bytes: bytes, filename: str, sheet_name: object = 0, header=None) -> pd.DataFrame:
    if is_csv_file(filename):
        return pd.read_csv(BytesIO(file_bytes), header=header)
    return read_excel_sheet(file_bytes, filename, sheet_name=sheet_name, header=header)


@st.cache_data(show_spinner=False)
def list_excel_sheets(file_bytes: bytes, filename: str) -> List[str]:
    engine = get_excel_engine(filename)
    if engine is None:
        raise ValueError(f"Unsupported Excel file type for `{filename}`.")
    with pd.ExcelFile(BytesIO(file_bytes), engine=engine) as workbook:
        return workbook.sheet_names


@st.cache_data(show_spinner=False)
def list_tabular_sheets(file_bytes: bytes, filename: str) -> List[str]:
    if is_csv_file(filename):
        return ["Sheet1"]
    return list_excel_sheets(file_bytes, filename)


def find_header_row(raw_df: pd.DataFrame) -> Optional[int]:
    required = {"FUNDCODE", "FUNDDESCRIPTION", "TOTALMVCAD", "MANDATECODE"}
    for idx, row in raw_df.iterrows():
        row_values = {normalize_header(val) for val in row.tolist() if normalize_text(val)}
        if required.issubset(row_values):
            return idx
    return None


def infer_saa_taa_column(df: pd.DataFrame) -> Optional[str]:
    exact = [col for col in df.columns if normalize_header(col) == "SAATAA"]
    if exact:
        return exact[0]

    for col in df.columns:
        values = (
            df[col]
            .dropna()
            .astype(str)
            .str.strip()
            .replace("", pd.NA)
            .dropna()
            .str.upper()
        )
        if not values.empty and set(values.unique()).issubset(set(HOLDING_TYPE_OPTIONS)):
            return col
    return None


def parse_holdings_file(file_bytes: bytes, filename: str) -> Tuple[pd.DataFrame, Dict[str, List[str]]]:
    raw_df = read_tabular_sheet(file_bytes, filename, sheet_name=0, header=None)
    sheet_names = list_tabular_sheets(file_bytes, filename)
    ips_sheet = next((name for name in sheet_names if normalize_key(name) == "IPS"), None)
    if ips_sheet is None:
        raise ValueError("The client holdings file must contain an `IPS` sheet.")

    raw_df = read_tabular_sheet(file_bytes, filename, sheet_name=ips_sheet, header=None)
    header_row = find_header_row(raw_df)
    if header_row is None:
        raise ValueError(
            "Unable to locate the IPS header row. Expected columns include `Fund Code`, `Fund Description`, `Total MV (CAD)`, and `mandate_code`."
        )

    header_values = [normalize_text(value) for value in raw_df.iloc[header_row].tolist()]
    data = raw_df.iloc[header_row + 1 :].copy()
    data.columns = [
        value if value else f"Unnamed_{idx}"
        for idx, value in enumerate(header_values)
    ]
    data = data.dropna(how="all").reset_index(drop=True)

    column_map: Dict[str, str] = {}
    for col in data.columns:
        normalized = normalize_header(col)
        if normalized == "FUNDCODE":
            column_map[col] = "Fund Code"
        elif normalized == "FUNDDESCRIPTION":
            column_map[col] = "Fund Description"
        elif normalized == "TOTALMVCAD":
            column_map[col] = "Total MV (CAD)"
        elif normalized == "MANDATECODE":
            column_map[col] = "mandate_code"
        elif normalized == "SAATAA":
            column_map[col] = "saa_taa"

    data = data.rename(columns=column_map)

    saa_taa_column = "saa_taa" if "saa_taa" in data.columns else None
    if saa_taa_column is None:
        inferred = infer_saa_taa_column(data)
        if inferred is None:
            raise ValueError("Unable to locate the `saa_taa` column in the IPS sheet.")
        data = data.rename(columns={inferred: "saa_taa"})

    required_columns = ["Fund Code", "Fund Description", "Total MV (CAD)", "mandate_code", "saa_taa"]
    missing_columns = [col for col in required_columns if col not in data.columns]
    if missing_columns:
        raise ValueError(f"Missing required IPS columns: {', '.join(missing_columns)}")

    holdings = data[required_columns].copy()
    holdings["Total MV (CAD)"] = coerce_number_series(holdings["Total MV (CAD)"])
    holdings["mandate_code"] = holdings["mandate_code"].apply(extract_mandate_code)
    holdings["saa_taa"] = holdings["saa_taa"].apply(normalize_holding_type)
    holdings["Fund Code"] = holdings["Fund Code"].astype(str).str.strip()
    holdings["Fund Description"] = holdings["Fund Description"].apply(normalize_text)
    holdings = holdings[holdings["Total MV (CAD)"].fillna(0) > 0].copy()
    holdings = holdings[holdings["mandate_code"].notna()].copy()
    holdings = holdings[holdings["saa_taa"].isin(HOLDING_TYPE_OPTIONS)].copy()
    holdings["allocation_bucket"] = holdings["saa_taa"].apply(allocation_bucket_for_holding_type)
    holdings = holdings.reset_index(drop=True)

    messages = {
        "warnings": [],
        "info": [],
    }

    if holdings.empty:
        raise ValueError("No IPS holdings rows with `Total MV (CAD)` greater than 0 were found.")

    duplicate_codes = holdings["mandate_code"][holdings["mandate_code"].duplicated()].unique().tolist()
    if duplicate_codes:
        messages["warnings"].append(
            f"Duplicate mandate codes detected in holdings and will reuse the same support file: {', '.join(sorted(duplicate_codes))}."
        )

    return holdings, messages


def derive_support_code_from_fund_code(fund_code: object) -> Optional[str]:
    text = normalize_text(fund_code)
    if not text.isdigit():
        return None

    numeric = int(text)
    derived = numeric - 2000
    if 10000 <= derived <= 99999:
        return str(derived)
    return None


def derive_class_support_code_from_factset_code(code: object) -> Optional[str]:
    normalized = normalize_code(code)
    if not normalized or not normalized.isdigit():
        return None
    numeric = int(normalized)
    if numeric % 2 != 0:
        return normalized
    candidate = numeric + 1
    if 10000 <= candidate <= 99999:
        return str(candidate)
    return None


def derive_250xx_support_code_from_fund_code(fund_code: object) -> Optional[str]:
    normalized = normalize_code(fund_code)
    if not normalized or len(normalized) != 5:
        return None
    suffix = normalized[-3:]
    candidate = f"25{suffix}"
    if candidate.isdigit():
        return candidate
    return None


def lookup_support_codes_from_factset(
    fund_code: object,
    holding_type: object = None,
    fund_description: object = None,
    factset_override: Optional[dict] = None,
) -> List[str]:
    entered_code = normalize_code(fund_code)
    if not entered_code:
        return []

    factset_table = get_factset_model_table(factset_override)
    if factset_table.empty or "sales_charge_code" not in factset_table.columns or "mandate_code" not in factset_table.columns:
        return []

    matches = factset_table[factset_table["sales_charge_code"] == entered_code].copy()
    normalized_holding_type = normalize_holding_type(holding_type)
    if normalized_holding_type in {"SAA", "TAA"} and "saa_taa" in matches.columns:
        typed_matches = matches[matches["saa_taa"] == normalized_holding_type].copy()
        if not typed_matches.empty:
            matches = typed_matches

    series_suffix = infer_model_series_suffix(fund_description)
    if series_suffix and "factset_model_code" in matches.columns:
        series_matches = matches[
            matches["factset_model_code"].apply(lambda value: normalize_text(value).upper().endswith(f"_{series_suffix}"))
        ].copy()
        if not series_matches.empty:
            matches = series_matches

    codes: List[str] = []
    for value in matches["mandate_code"].tolist():
        code = normalize_code(value)
        if code and code not in codes:
            codes.append(code)
    return codes


def build_support_candidates(
    fund_code: object,
    mandate_code: object = None,
    holding_type: object = None,
    fund_description: object = None,
    factset_override: Optional[dict] = None,
) -> List[str]:
    candidates: List[str] = []
    entered = extract_mandate_code(mandate_code)
    derived = derive_support_code_from_fund_code(fund_code)
    factset_codes = lookup_support_codes_from_factset(
        fund_code,
        holding_type=holding_type,
        fund_description=fund_description,
        factset_override=factset_override,
    )

    series_suffix = infer_model_series_suffix(fund_description)
    if series_suffix == "C":
        if derived:
            candidates.append(derived)
        direct_class_candidate = derive_250xx_support_code_from_fund_code(fund_code)
        if direct_class_candidate:
            candidates.append(direct_class_candidate)
        for factset_code in factset_codes:
            class_candidate = derive_class_support_code_from_factset_code(factset_code)
            if class_candidate:
                candidates.append(class_candidate)
        if entered:
            candidates.append(entered)
            entered_class_candidate = derive_class_support_code_from_factset_code(entered)
            if entered_class_candidate:
                candidates.append(entered_class_candidate)
    else:
        candidates.extend(factset_codes)
        if derived:
            candidates.append(derived)

    if entered and series_suffix != "C":
        candidates.append(entered)
        if entered.isdigit():
            candidates.append(str(int(entered) + 1).zfill(len(entered)))
            if int(entered) > 0:
                candidates.append(str(int(entered) - 1).zfill(len(entered)))

    deduped: List[str] = []
    for code in candidates:
        if code and code not in deduped:
            deduped.append(code)
    return deduped


def lookup_sma_row(fund_code: object, sma_override: Optional[dict] = None) -> Optional[pd.Series]:
    code = normalize_code(fund_code)
    if not code:
        return None
    sma_table = get_sma_grouping_table(sma_override)
    if sma_table.empty:
        return None
    matches = sma_table[sma_table["Fund Code"] == code]
    if matches.empty:
        return None
    return matches.iloc[0]


def infer_sma_row_from_description(fund_code: object, description: object) -> Optional[pd.Series]:
    text = normalize_key(description)
    if "PRIVATE MARKETS" in text or "PRIVATE MARKET" in text:
        return pd.Series(
            {
                "Fund Code": normalize_code(fund_code) or normalize_text(fund_code),
                "Sales Charge Code.Legal Name": normalize_text(description),
                "Portfolio Composition": "Private Alt",
                "Portfolio Breakdown": "Alternatives",
                "Portfolio AAbA": "Alternatives",
            }
        )
    if "HIGH INTEREST SAVINGS" in text:
        return pd.Series(
            {
                "Fund Code": normalize_code(fund_code) or normalize_text(fund_code),
                "Sales Charge Code.Legal Name": normalize_text(description),
                "Portfolio Composition": "Cash",
                "Portfolio Breakdown": "Cash",
                "Portfolio AAbA": "Income (Incl. Cash)",
            }
        )
    if "ENHANCED SHORT DURATION BOND" in text:
        return pd.Series(
            {
                "Fund Code": normalize_code(fund_code) or normalize_text(fund_code),
                "Sales Charge Code.Legal Name": normalize_text(description),
                "Portfolio Composition": "Income",
                "Portfolio Breakdown": "Income",
                "Portfolio AAbA": "Income (Incl. Cash)",
            }
        )
    if "SELECT INCOME MANAGED" in text:
        return pd.Series(
            {
                "Fund Code": normalize_code(fund_code) or normalize_text(fund_code),
                "Sales Charge Code.Legal Name": normalize_text(description),
                "Portfolio Composition": "Balanced",
                "Portfolio Breakdown": "Income",
                "Portfolio AAbA": "Balanced",
            }
        )
    return None


def get_sma_row_for_holding(holding: pd.Series, sma_override: Optional[dict] = None) -> Optional[pd.Series]:
    if normalize_holding_type(holding.get("saa_taa")) != "SMA":
        return None
    row = lookup_sma_row(holding.get("Fund Code"), sma_override=sma_override)
    if row is not None:
        return row
    return infer_sma_row_from_description(
        holding.get("Fund Code"),
        holding.get("Fund Description"),
    )


def map_sma_composition_group(portfolio_composition: object, portfolio_breakdown: object) -> str:
    composition = normalize_key(portfolio_composition)

    if composition == "INCOME":
        return "Income"
    if composition == "EQUITY":
        return "Equity"
    if composition == "BALANCED":
        return "Balanced"
    if composition == "LIQUID ALT":
        return "Liquid Alternatives"
    if composition == "SECTOR":
        return "Sector"
    if composition == "CASH":
        return "Cash"
    if composition in {"PRIVATE ALT", "PRIVATE ALTERNATIVES"}:
        return "Private Alternatives"
    if composition == "OTHER":
        return "Other"
    return ""


def map_sma_breakdown_group(portfolio_breakdown: object) -> str:
    breakdown = normalize_key(portfolio_breakdown)
    return {
        "INCOME": "Income",
        "INTERNATIONAL EQUITY": "International Equity",
        "US EQUITY": "US Equity",
        "CANADIAN EQUITY": "Canadian Equity",
        "CASH": "Cash",
        "OTHER": "Other",
        "ALTERNATIVES": "Alternatives",
        "PRIVATE ALT": "Alternatives",
        "PRIVATE ALTERNATIVES": "Alternatives",
        "LIQUID ALT": "Alternatives",
    }.get(breakdown, "")


def build_sma_rows(holding: pd.Series, sma_row: pd.Series) -> pd.DataFrame:
    holding_mv = float(holding["Total MV (CAD)"])
    fund_code = normalize_code(holding["Fund Code"]) or normalize_text(holding["Fund Code"])
    component_name = normalize_text(sma_row.get("Sales Charge Code.Legal Name")) or holding["Fund Description"]
    composition_group = map_sma_composition_group(
        sma_row.get("Portfolio Composition"),
        sma_row.get("Portfolio Breakdown"),
    )
    breakdown_group = map_sma_breakdown_group(sma_row.get("Portfolio Breakdown"))

    return pd.DataFrame(
        [
            {
                "Component": component_name,
                "Port. Weight": 100.0,
                "Fund Code": holding["Fund Code"],
                "Fund Description": holding["Fund Description"],
                "Total MV (CAD)": holding_mv,
                "mandate_code": holding.get("mandate_code"),
                "support_code": fund_code,
                "saa_taa": holding["allocation_bucket"],
                "Holding Type": holding["saa_taa"],
                "Support File": "SMA Reference",
                "Block Label": f"{holding['Fund Description']} - SMA - {holding['saa_taa']}",
                "Weighted MV (CAD)": holding_mv,
                "Composition Group": composition_group,
                "Breakdown Group": breakdown_group,
                "Source Type": "SMA",
            }
        ]
    )


def detect_factset_models(
    holdings: pd.DataFrame,
    factset_override: Optional[dict] = None,
) -> Tuple[List[dict], List[str], List[str]]:
    info: List[str] = []
    warnings: List[str] = []
    factset_table = get_factset_model_table(factset_override)
    if factset_table.empty:
        return [], warnings, info

    taa_holdings = holdings[holdings["saa_taa"] == "TAA"].copy()
    if taa_holdings.empty:
        return [], warnings, info

    model_records: List[dict] = []
    for _, holding in taa_holdings.iterrows():
        lookup_codes = build_support_candidates(
            holding["Fund Code"],
            holding.get("mandate_code"),
            fund_description=holding.get("Fund Description"),
            factset_override=factset_override,
        )
        entered_code = normalize_code(holding["Fund Code"])
        matches = factset_table[factset_table["sales_charge_code"] == entered_code].copy()
        if matches.empty:
            matches = factset_table[factset_table["sales_charge_code"].isin(lookup_codes)].copy()
        series_suffix = infer_model_series_suffix(holding.get("Fund Description"))
        if series_suffix and "factset_model_code" in matches.columns:
            series_matches = matches[
                matches["factset_model_code"].apply(lambda value: normalize_text(value).upper().endswith(f"_{series_suffix}"))
            ].copy()
            if not series_matches.empty:
                matches = series_matches
        unique_models = sorted(model for model in matches["factset_model_code"].dropna().unique().tolist() if model)
        if not unique_models:
            warnings.append(
                f"No FactSet model mapping was found for TAA holding `{holding['Fund Description']}`."
            )
            continue
        if len(unique_models) > 1:
            warnings.append(
                f"Multiple FactSet models were found for TAA holding `{holding['Fund Description']}`: {', '.join(unique_models)}."
            )

        for model_code in unique_models:
            base_code, _, suffix = model_code.partition("_")
            model_records.append(
                {
                    "Fund Code": holding["Fund Code"],
                    "Fund Description": holding["Fund Description"],
                    "FactSet Model": model_code,
                    "Model Allocation": base_code,
                    "Model Series": suffix or "",
                }
            )

    unique_model_codes = sorted({record["FactSet Model"] for record in model_records})
    if len(unique_model_codes) == 1:
        info.append(f"Detected FactSet model: {unique_model_codes[0]}.")
    elif len(unique_model_codes) > 1:
        warnings.append(
            "Multiple FactSet models were detected across the entered TAA holdings: "
            + ", ".join(unique_model_codes)
            + "."
        )

    return model_records, warnings, info


def load_draft_state() -> Tuple[pd.DataFrame, List[dict], Optional[dict], Optional[dict]]:
    if not DRAFT_PATH.exists():
        return DEFAULT_HOLDINGS_INPUT.copy(), [], None, None

    try:
        payload = json.loads(DRAFT_PATH.read_text())
    except Exception:
        return DEFAULT_HOLDINGS_INPUT.copy(), [], None, None

    holdings_records = payload.get("holdings", [])
    holdings_text = payload.get("holdings_text")
    support_records = payload.get("support_files", [])
    sma_override_record = payload.get("sma_override_file")
    factset_override_record = payload.get("factset_model_file")

    if holdings_text:
        try:
            holdings_df = parse_holdings_text(holdings_text)
        except Exception:
            holdings_df = pd.DataFrame(holdings_records) if holdings_records else DEFAULT_HOLDINGS_INPUT.copy()
    else:
        holdings_df = pd.DataFrame(holdings_records) if holdings_records else DEFAULT_HOLDINGS_INPUT.copy()
    for column in MANUAL_HOLDINGS_COLUMNS:
        if column not in holdings_df.columns:
            holdings_df[column] = DEFAULT_HOLDINGS_INPUT.iloc[0].get(column, "")
    holdings_df = holdings_df[MANUAL_HOLDINGS_COLUMNS].copy()

    saved_support_files: List[dict] = []
    for record in support_records:
        decoded = decode_saved_record(record)
        if decoded:
            saved_support_files.append(decoded)

    return (
        holdings_df,
        saved_support_files,
        decode_saved_record(sma_override_record),
        decode_saved_record(factset_override_record),
    )


def save_draft_state(
    holdings_df: pd.DataFrame,
    saved_support_files: List[dict],
    holdings_text: str,
    sma_override_file: Optional[dict],
    factset_model_file: Optional[dict],
) -> None:
    payload = {
        "holdings": holdings_df.where(pd.notnull(holdings_df), None).to_dict("records"),
        "holdings_text": holdings_text,
        "support_files": [
            encode_saved_record(item)
            for item in saved_support_files
            if encode_saved_record(item)
        ],
        "sma_override_file": encode_saved_record(sma_override_file),
        "factset_model_file": encode_saved_record(factset_model_file),
    }
    DRAFT_PATH.write_text(json.dumps(payload))


def clear_draft_state() -> None:
    if DRAFT_PATH.exists():
        DRAFT_PATH.unlink()


def load_account_history() -> List[dict]:
    if not HISTORY_PATH.exists():
        return []
    try:
        payload = json.loads(HISTORY_PATH.read_text())
    except Exception:
        return []
    entries = payload.get("entries", [])
    if not isinstance(entries, list):
        return []
    return entries


def save_account_history(entries: List[dict]) -> None:
    safe_entries = entries[:HISTORY_LIMIT]
    HISTORY_PATH.write_text(json.dumps({"entries": safe_entries}))


def format_saved_datetime(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return "No saved date"
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return text
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=APP_TIMEZONE)
    local_dt = parsed.astimezone(APP_TIMEZONE)
    date_text = local_dt.strftime("%b %d, %Y")
    time_text = local_dt.strftime("%I:%M %p").lstrip("0")
    return f"{date_text} {time_text}"


def format_history_entry(entry: dict) -> str:
    label = normalize_text(entry.get("label")) or "Untitled portfolio"
    total_mv = entry.get("portfolio_total")
    total_text = "Unknown MV"
    if isinstance(total_mv, (int, float)):
        total_text = format_currency(float(total_mv))
    date_text = format_saved_datetime(entry.get("created_at"))
    holdings_count = len(entry.get("holdings", []) or [])
    return f"{date_text} | {total_text} | {label} | {holdings_count} holdings"


def format_history_entry_details(entry: dict) -> str:
    label = normalize_text(entry.get("label")) or "Untitled portfolio"
    total_mv = entry.get("portfolio_total")
    total_text = format_currency(float(total_mv)) if isinstance(total_mv, (int, float)) else "Unknown"
    holdings = entry.get("holdings", []) or []
    support_files = entry.get("support_files", []) or []
    models = []
    for model in entry.get("factset_models", []) or []:
        model_name = normalize_text(model.get("FactSet Model") if isinstance(model, dict) else model)
        if model_name:
            models.append(model_name)
    model_text = ", ".join(models) if models else "No model detected"
    date_text = format_saved_datetime(entry.get("created_at"))
    codes = [
        normalize_text(row.get("Fund Code"))
        for row in holdings
        if isinstance(row, dict) and normalize_text(row.get("Fund Code"))
    ]
    codes_text = ", ".join(codes[:8])
    if len(codes) > 8:
        codes_text += f", +{len(codes) - 8} more"
    return (
        f"Saved: {date_text}\n\n"
        f"Label: {label}\n\n"
        f"Market value: {total_text}\n\n"
        f"Model: {model_text}\n\n"
        f"Holdings: {len(holdings)} ({codes_text or 'none'})\n\n"
        f"Files: {len(support_files)}"
    )


def build_default_history_label(results: dict) -> str:
    factset_models = results.get("factset_models", [])
    if factset_models:
        model = normalize_text(factset_models[0].get("FactSet Model"))
        if model:
            return f"{model} portfolio"
    holdings = results.get("holdings")
    if isinstance(holdings, pd.DataFrame) and not holdings.empty:
        fund_codes = ", ".join(holdings["Fund Code"].astype(str).head(3).tolist())
        return f"Portfolio {fund_codes}"
    return "Portfolio calculation"


def save_account_history_entry(
    label: str,
    holdings_df: pd.DataFrame,
    saved_support_files: List[dict],
    sma_override_file: Optional[dict],
    factset_model_file: Optional[dict],
    results: dict,
) -> dict:
    created_at = datetime.now(APP_TIMEZONE).isoformat(timespec="seconds")
    clean_label = normalize_text(label) or build_default_history_label(results)
    entry = {
        "id": uuid4().hex,
        "created_at": created_at,
        "label": clean_label,
        "portfolio_total": float(results.get("portfolio_total", 0.0)),
        "factset_models": results.get("factset_models", []),
        "holdings": holdings_df.where(pd.notnull(holdings_df), None).to_dict("records"),
        "support_files": [
            encoded
            for encoded in (encode_saved_record(item) for item in saved_support_files)
            if encoded
        ],
        "sma_override_file": encode_saved_record(sma_override_file),
        "factset_model_file": encode_saved_record(factset_model_file),
    }
    entries = load_account_history()
    entries = [existing for existing in entries if existing.get("id") != entry["id"]]
    entries.insert(0, entry)
    save_account_history(entries)
    return entry


def load_history_entry_into_session(entry: dict) -> None:
    holdings_records = entry.get("holdings", [])
    holdings_df = pd.DataFrame(holdings_records) if holdings_records else DEFAULT_HOLDINGS_INPUT.copy()
    for column in MANUAL_HOLDINGS_COLUMNS:
        if column not in holdings_df.columns:
            holdings_df[column] = DEFAULT_HOLDINGS_INPUT.iloc[0].get(column, "")
    holdings_df = holdings_df[MANUAL_HOLDINGS_COLUMNS].copy()
    support_files = [
        decoded
        for decoded in (decode_saved_record(record) for record in entry.get("support_files", []))
        if decoded
    ]
    st.session_state["holdings_rows"] = pad_holding_rows(holdings_df.to_dict("records"))
    st.session_state["holdings_paste_text"] = holdings_df_to_text(holdings_df)
    st.session_state["saved_support_files"] = support_files
    st.session_state["saved_sma_override_file"] = decode_saved_record(entry.get("sma_override_file"))
    st.session_state["saved_factset_model_file"] = decode_saved_record(entry.get("factset_model_file"))
    st.session_state["account_label"] = normalize_text(entry.get("label"))
    st.session_state["pending_history_recalculate"] = True
    st.session_state["widget_reset_nonce"] = st.session_state.get("widget_reset_nonce", 0) + 1
    save_draft_state(
        strip_blank_holding_rows(holdings_df),
        support_files,
        st.session_state["holdings_paste_text"],
        st.session_state.get("saved_sma_override_file"),
        st.session_state.get("saved_factset_model_file"),
    )
    clear_latest_calculation()


def delete_history_entry(entry_id: str) -> None:
    entries = [entry for entry in load_account_history() if entry.get("id") != entry_id]
    save_account_history(entries)


def queue_history_entry_load(entry_id: str) -> None:
    st.session_state["pending_history_load_id"] = entry_id


def apply_queued_history_entry_load() -> None:
    entry_id = st.session_state.pop("pending_history_load_id", None)
    if not entry_id:
        return
    for entry in load_account_history():
        if entry.get("id") == entry_id:
            load_history_entry_into_session(entry)
            return


def blank_holding_row() -> dict:
    return {
        "Fund Code": "",
        "Fund Description": "",
        "Total MV (CAD)": "",
        "saa_taa": "SAA",
    }


def pad_holding_rows(rows: List[dict], minimum: int = 5) -> List[dict]:
    padded = list(rows)
    while len(padded) < minimum:
        padded.append(blank_holding_row())
    return padded


def holding_rows_to_df(rows: List[dict]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame(columns=MANUAL_HOLDINGS_COLUMNS)
    return pd.DataFrame(rows)[MANUAL_HOLDINGS_COLUMNS].copy()


def strip_blank_holding_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=MANUAL_HOLDINGS_COLUMNS)
    cleaned = df.copy()
    for column in MANUAL_HOLDINGS_COLUMNS:
        if column not in cleaned.columns:
            cleaned[column] = ""
    cleaned = cleaned[MANUAL_HOLDINGS_COLUMNS].copy()
    normalized_fund_code = cleaned["Fund Code"].apply(normalize_text)
    normalized_description = cleaned["Fund Description"].apply(normalize_text)
    normalized_total_mv = cleaned["Total MV (CAD)"].apply(normalize_text)
    blank_mask = (
        normalized_fund_code.eq("")
        & normalized_description.eq("")
        & normalized_total_mv.eq("")
    )
    return cleaned.loc[~blank_mask].reset_index(drop=True)


def holdings_df_to_text(df: pd.DataFrame) -> str:
    export_df = strip_blank_holding_rows(df)
    if export_df.empty:
        return HOLDINGS_TEXT_TEMPLATE
    return export_df.to_csv(sep="\t", index=False)


def clean_holdings_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    cleaned = df.copy()
    for column in MANUAL_HOLDINGS_COLUMNS:
        if column not in cleaned.columns:
            cleaned[column] = ""
    cleaned = cleaned[MANUAL_HOLDINGS_COLUMNS].copy()
    cleaned = cleaned.where(pd.notnull(cleaned), "")
    cleaned["Fund Code"] = cleaned["Fund Code"].apply(
        lambda value: normalize_text(value) if normalize_text(value).lower() != "nan" else ""
    )
    cleaned["Fund Description"] = cleaned["Fund Description"].apply(
        lambda value: normalize_text(value) if normalize_text(value).lower() != "nan" else ""
    )
    cleaned["Total MV (CAD)"] = cleaned["Total MV (CAD)"].apply(
        lambda value: "" if normalize_text(value).lower() in {"", "nan"} else normalize_text(value)
    )
    cleaned["saa_taa"] = cleaned["saa_taa"].apply(lambda value: normalize_holding_type(value) or "SAA")
    return cleaned


def infer_holding_type_from_description(description: object) -> str:
    text = normalize_key(description)
    if "TACTICAL ASSET ALLOCATION" in text or " TAA " in f" {text} ":
        return "TAA"
    if " SMA " in f" {text} ":
        return "SMA"
    return "SAA"


def is_managed_portfolio_support_holding(description: object) -> bool:
    text = f" {normalize_key(description)} "
    if " TACTICAL ASSET ALLOCATION " in text:
        return False
    if " APP " in text:
        return False
    return " MANAGED POOL " in text or " MANAGED CLASS " in text


def apply_sma_type_detection(df: pd.DataFrame, sma_override: Optional[dict] = None) -> pd.DataFrame:
    """Mark pasted holdings as SMA when their fund code exists in the SMA reference."""
    if df is None or df.empty:
        return pd.DataFrame(columns=MANUAL_HOLDINGS_COLUMNS)

    detected = clean_holdings_dataframe(df)
    try:
        sma_table = get_sma_grouping_table(sma_override)
    except Exception:
        return detected

    if sma_table.empty or "Fund Code" not in sma_table.columns:
        return detected

    sma_codes = {
        code
        for code in sma_table["Fund Code"].apply(normalize_code).tolist()
        if code
    }
    if not sma_codes:
        return detected

    fund_codes = detected["Fund Code"].apply(normalize_code)
    current_types = detected["saa_taa"].apply(normalize_holding_type)
    managed_portfolio_mask = detected["Fund Description"].apply(is_managed_portfolio_support_holding)
    sma_mask = fund_codes.isin(sma_codes) & current_types.ne("TAA") & ~managed_portfolio_mask
    detected.loc[sma_mask, "saa_taa"] = "SMA"
    return detected


def parse_holdings_tabular_export(holdings_text: str) -> Optional[pd.DataFrame]:
    try:
        table = pd.read_csv(
            StringIO(holdings_text),
            sep="\t",
            header=None,
            dtype=str,
            engine="python",
            keep_default_na=False,
        )
    except Exception:
        return None
    if table.empty or table.shape[1] < 2:
        return None

    header_idx: Optional[int] = None
    code_col: Optional[int] = None
    name_col: Optional[int] = None
    mv_col: Optional[int] = None
    type_col: Optional[int] = None
    for row_idx, row in table.iterrows():
        normalized_cells = [normalize_header(value) for value in row.tolist()]
        if "FUNDCODE" not in normalized_cells:
            continue
        possible_code_col = normalized_cells.index("FUNDCODE")
        possible_name_col = next(
            (
                idx
                for idx, value in enumerate(normalized_cells)
                if value in {"FUNDNAME", "FUNDDESCRIPTION", "FUNDDESC"}
            ),
            None,
        )
        possible_mv_col = next(
            (
                idx
                for idx, value in enumerate(normalized_cells)
                if value in {
                    "TOTALMVCAD",
                    "MARKETVALUECDN",
                    "MARKETVALUECDNS",
                    "MARKETVALUECAD",
                    "MARKETVALUECADS",
                    "MARKETVALUECDN",
                    "MARKETVALUEC",
                    "MARKETVALUECA",
                    "MARKETVALUECADOLLAR",
                    "MARKETVALUECDNDOLLAR",
                }
            ),
            None,
        )
        if possible_mv_col is None:
            generic_mv_cols = [
                idx
                for idx, value in enumerate(normalized_cells)
                if value in {"MARKETVALUE", "TOTALMV", "MARKETVAL"}
            ]
            if generic_mv_cols:
                possible_mv_col = generic_mv_cols[-1]
        possible_type_col = next(
            (
                idx
                for idx, value in enumerate(normalized_cells)
                if value in {"SAATAA", "TYPE", "HOLDINGTYPE"}
            ),
            None,
        )
        if possible_name_col is not None and possible_mv_col is not None:
            header_idx = row_idx
            code_col = possible_code_col
            name_col = possible_name_col
            mv_col = possible_mv_col
            type_col = possible_type_col
            break

    if header_idx is None or code_col is None or name_col is None or mv_col is None:
        return None

    records: List[dict] = []
    for _, row in table.iloc[header_idx + 1 :].iterrows():
        cells = row.tolist()
        line_text = " ".join(normalize_text(value) for value in cells if normalize_text(value))
        if normalize_key(line_text).startswith("TOTAL "):
            break

        code = normalize_code(cells[code_col] if code_col < len(cells) else "")
        description = normalize_text(cells[name_col] if name_col < len(cells) else "")
        market_value = normalize_text(cells[mv_col] if mv_col < len(cells) else "")
        explicit_type = normalize_holding_type(cells[type_col]) if type_col is not None and type_col < len(cells) else ""
        if not code or not description or not market_value:
            continue
        records.append(
            {
                "Fund Code": code,
                "Fund Description": description,
                "Total MV (CAD)": market_value,
                "saa_taa": explicit_type or infer_holding_type_from_description(description),
            }
        )

    if not records:
        return None
    return pd.DataFrame(records, columns=MANUAL_HOLDINGS_COLUMNS)


def parse_holdings_line_export(holdings_text: str) -> Optional[pd.DataFrame]:
    lines = [normalize_text(line) for line in holdings_text.splitlines()]
    lines = [line for line in lines if line]
    if not lines:
        return None

    ignored_headers = {
        "FUND CODE",
        "FUND NAME",
        "MARKET VALUE US$",
        "MARKET VALUE USD",
        "EXCHANGE RATE",
        "MARKET VALUE CDN$",
        "MARKET VALUE CAD$",
        "MARKET VALUE (CAD)",
        "MARKET VALUE CAD",
        "MARKET VALUE C$",
        "MARKET VALUE CA$",
        "MARKET VALUE CDN",
        "MARKET VALUE CAD",
    }
    stop_prefixes = ("TOTAL ",)

    records: List[dict] = []
    idx = 0
    while idx < len(lines):
        line = lines[idx]
        normalized = normalize_key(line)

        if normalized in ignored_headers:
            idx += 1
            continue
        if normalized.startswith(stop_prefixes):
            idx += 1
            continue

        same_line_match = re.match(
            r"^(?P<code>\d{4,6})\s+(?P<description>.+?)\s+(?P<market_value>[$€£]?\s*-?[\d,]+(?:\.\d+)?)\s*(?P<holding_type>SAA|TAA|SMA)?$",
            line,
            flags=re.IGNORECASE,
        )
        if same_line_match:
            description = normalize_text(same_line_match.group("description"))
            market_value = normalize_text(same_line_match.group("market_value"))
            explicit_type = normalize_holding_type(same_line_match.group("holding_type"))
            records.append(
                {
                    "Fund Code": normalize_code(same_line_match.group("code")) or same_line_match.group("code"),
                    "Fund Description": description,
                    "Total MV (CAD)": market_value,
                    "saa_taa": explicit_type or infer_holding_type_from_description(description),
                }
            )
            idx += 1
            continue

        code = normalize_code(line) if re.fullmatch(r"\d{4,6}", line) else None
        if code is None:
            idx += 1
            continue

        description = ""
        market_values: List[str] = []
        cursor = idx + 1
        while cursor < len(lines):
            candidate = lines[cursor]
            candidate_normalized = normalize_key(candidate)
            if candidate_normalized in ignored_headers:
                cursor += 1
                continue
            if candidate_normalized.startswith(stop_prefixes):
                break
            if extract_mandate_code(candidate) is not None or re.fullmatch(r"\d{4,6}", candidate):
                break
            if description == "":
                description = candidate
                cursor += 1
                continue
            if re.search(r"[$€£]?\s*-?[\d,]+(?:\.\d+)?", candidate):
                if not candidate.strip().startswith("@"):
                    market_values.append(candidate)
                cursor += 1
                continue
            cursor += 1

        market_value = market_values[-1] if market_values else ""
        if description and market_value:
            records.append(
                {
                    "Fund Code": code,
                    "Fund Description": description,
                    "Total MV (CAD)": market_value,
                    "saa_taa": infer_holding_type_from_description(description),
                }
            )
            idx = cursor
            continue

        idx += 1

    if not records:
        return None
    return pd.DataFrame(records, columns=MANUAL_HOLDINGS_COLUMNS)


def parse_holdings_text(holdings_text: str) -> pd.DataFrame:
    text = holdings_text.strip()
    if not text:
        return pd.DataFrame(columns=MANUAL_HOLDINGS_COLUMNS)

    tabular_export = parse_holdings_tabular_export(text)
    if tabular_export is not None:
        return clean_holdings_dataframe(tabular_export)

    line_export = parse_holdings_line_export(text)
    if line_export is not None:
        return clean_holdings_dataframe(line_export)

    try:
        parsed = pd.read_csv(StringIO(text), sep=None, engine="python")
        normalized_columns = {col: normalize_header(col) for col in parsed.columns}

        if set(normalized_columns.values()) >= {"FUNDCODE", "FUNDDESCRIPTION", "TOTALMVCAD", "SAATAA"}:
            rename_map = {}
            for col, normalized in normalized_columns.items():
                if normalized == "FUNDCODE":
                    rename_map[col] = "Fund Code"
                elif normalized == "FUNDDESCRIPTION":
                    rename_map[col] = "Fund Description"
                elif normalized == "TOTALMVCAD":
                    rename_map[col] = "Total MV (CAD)"
                elif normalized == "SAATAA":
                    rename_map[col] = "saa_taa"
            parsed = parsed.rename(columns=rename_map)
            return clean_holdings_dataframe(parsed)
    except Exception:
        pass

    parsed = pd.read_csv(
        StringIO(text),
        sep=None,
        engine="python",
        header=None,
        names=MANUAL_HOLDINGS_COLUMNS,
    )
    return clean_holdings_dataframe(parsed)


def parse_manual_holdings_input(
    input_df: pd.DataFrame,
    sma_override: Optional[dict] = None,
    factset_override: Optional[dict] = None,
) -> Tuple[pd.DataFrame, Dict[str, List[str]]]:
    if input_df is None:
        raise ValueError("Enter at least one holdings row before running the calculation.")

    holdings = input_df.copy()
    if "mandate_code" not in holdings.columns:
        holdings["mandate_code"] = ""

    required_columns = MANUAL_HOLDINGS_COLUMNS + ["mandate_code"]
    missing_columns = [col for col in MANUAL_HOLDINGS_COLUMNS if col not in holdings.columns]
    if missing_columns:
        raise ValueError(f"Missing required holdings columns: {', '.join(missing_columns)}")

    holdings = holdings[required_columns].copy()
    holdings["Fund Code"] = holdings["Fund Code"].apply(normalize_text)
    holdings["Fund Description"] = holdings["Fund Description"].apply(normalize_text)
    holdings["Total MV (CAD)"] = coerce_number_series(holdings["Total MV (CAD)"])
    holdings["mandate_code"] = holdings["mandate_code"].apply(extract_mandate_code)
    holdings["saa_taa"] = holdings["saa_taa"].apply(normalize_holding_type)

    messages = {
        "warnings": [],
        "info": [],
    }

    all_blank_mask = (
        (holdings["Fund Code"] == "")
        & (holdings["Fund Description"] == "")
        & (holdings["Total MV (CAD)"].isna())
    )
    blank_rows = int(all_blank_mask.sum())
    if blank_rows:
        messages["info"].append(f"Ignored {blank_rows} blank holdings row(s).")
    holdings = holdings[~all_blank_mask].copy()

    if holdings.empty:
        raise ValueError("Enter at least one complete holdings row before running the calculation.")

    invalid_saa_taa = holdings[~holdings["saa_taa"].isin(HOLDING_TYPE_OPTIONS)]
    if not invalid_saa_taa.empty:
        raise ValueError("Each holdings row must have `saa_taa` set to `SAA`, `TAA`, or `SMA`.")

    missing_required = holdings[
        (holdings["Fund Code"] == "")
        | (holdings["Fund Description"] == "")
        | (holdings["Total MV (CAD)"].isna())
    ]
    if not missing_required.empty:
        raise ValueError(
            "Each holdings row must include Fund Code, Fund Description, Total MV (CAD), and saa_taa."
        )

    zero_mv_rows = holdings[holdings["Total MV (CAD)"].fillna(0) <= 0]
    if not zero_mv_rows.empty:
        messages["warnings"].append(
            f"Ignored {len(zero_mv_rows)} holdings row(s) with Total MV (CAD) less than or equal to 0."
        )
    holdings = holdings[holdings["Total MV (CAD)"].fillna(0) > 0].copy()

    if holdings.empty:
        raise ValueError("All entered holdings rows have Total MV (CAD) less than or equal to 0.")

    holdings["allocation_bucket"] = holdings["saa_taa"].apply(allocation_bucket_for_holding_type)
    holdings["support_code"] = holdings.apply(
        lambda row: (
            build_support_candidates(
                row["Fund Code"],
                holding_type=row["saa_taa"],
                fund_description=row["Fund Description"],
                factset_override=factset_override,
            )[0]
            if build_support_candidates(
                row["Fund Code"],
                holding_type=row["saa_taa"],
                fund_description=row["Fund Description"],
                factset_override=factset_override,
            )
            else None
        ),
        axis=1,
    )
    holdings["has_sma_mapping"] = holdings.apply(
        lambda row: pd.isna(row["support_code"]) and get_sma_row_for_holding(row, sma_override=sma_override) is not None,
        axis=1,
    )
    unresolved_support_rows = holdings[holdings["support_code"].isna() & ~holdings["has_sma_mapping"]]
    if not unresolved_support_rows.empty:
        raise ValueError(
            "Unable to derive the expected support file code from one or more Fund Code values."
        )

    sma_count = int(holdings["has_sma_mapping"].sum())
    if sma_count:
        messages["info"].append(
            f"Detected {sma_count} holding(s) with SMA fallback mapping. Those rows can be calculated from the bundled SMA asset-class reference when no support file is provided."
        )

    holdings = holdings.reset_index(drop=True)
    return holdings, messages


def parse_manual_holdings_text(input_text: str) -> Tuple[pd.DataFrame, Dict[str, List[str]]]:
    parsed_df = parse_holdings_text(input_text)
    return parse_manual_holdings_input(parsed_df)


def find_match_position(df: pd.DataFrame, target: str) -> Optional[Tuple[int, int]]:
    target_key = normalize_key(target)
    for row_idx in range(df.shape[0]):
        for col_idx in range(df.shape[1]):
            if normalize_key(df.iat[row_idx, col_idx]) == target_key:
                return row_idx, col_idx
    return None


def parse_support_report_date(raw_df: pd.DataFrame, anchor_row: int) -> Optional[pd.Timestamp]:
    window_start = max(anchor_row - 12, 0)
    date_pattern = re.compile(r"\b\d{1,2}[-/][A-Z]{3}[-/]\d{2,4}\b|\b\d{4}-\d{2}-\d{2}\b", re.IGNORECASE)

    for row_idx in range(anchor_row, window_start - 1, -1):
        for value in raw_df.iloc[row_idx].tolist():
            if pd.isna(value):
                continue
            if isinstance(value, pd.Timestamp):
                return value
            text = normalize_text(value)
            if not text:
                continue
            match = date_pattern.search(text)
            if not match:
                continue
            parsed = pd.to_datetime(match.group(0), errors="coerce", dayfirst=True)
            if pd.notna(parsed):
                return parsed

    return None


def format_reporting_period(report_date: Optional[pd.Timestamp]) -> Optional[str]:
    if report_date is None or pd.isna(report_date):
        return None
    return pd.Timestamp(report_date).strftime("%B %Y")


def build_reporting_period_note(reporting_period: Optional[str]) -> Optional[str]:
    if not reporting_period:
        return None
    return f"Asset Allocation sector weight is reported as of {reporting_period}."


def read_excel_hierarchy_metadata(
    file_bytes: bytes,
    filename: str,
    start_row_zero_based: int,
    name_col_zero_based: int,
) -> pd.DataFrame:
    if filename.lower().endswith(".csv") or get_excel_engine(filename) != "openpyxl":
        return pd.DataFrame(columns=["Sheet Row Number", "Hierarchy Indent", "Is Summary Row"])

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    records: List[dict] = []
    excel_col = name_col_zero_based + 1

    for excel_row in range(start_row_zero_based + 1, worksheet.max_row + 1):
        cell = worksheet.cell(excel_row, excel_col)
        indent = cell.alignment.indent if cell.alignment else None
        records.append(
            {
                "Sheet Row Number": excel_row,
                "Hierarchy Indent": float(indent) if indent is not None else pd.NA,
                "Is Summary Row": bool(cell.font.bold),
            }
        )

    return pd.DataFrame(records)


def select_support_rows_by_hierarchy(comp_df: pd.DataFrame, section: str) -> pd.Series:
    support_mask = comp_df["Source Type"].eq("Support File")
    if not support_mask.any():
        return pd.Series(True, index=comp_df.index)

    if "Hierarchy Indent" not in comp_df.columns:
        return pd.Series(True, index=comp_df.index)

    indent = pd.to_numeric(comp_df["Hierarchy Indent"], errors="coerce")
    has_hierarchy = indent.notna()
    keep = pd.Series(True, index=comp_df.index)

    if section in {"composition", "breakdown"}:
        normalized = comp_df["Component"].map(normalize_key)
        if section == "composition":
            support_keep = (~has_hierarchy) | indent.eq(4.0)
            support_keep |= comp_df["Composition Mapping Source"].eq("Funds/Alternatives Detection")
        else:
            support_keep = pd.Series(False, index=comp_df.index)
            support_keep |= ~has_hierarchy
            support_keep |= indent.eq(4.0) & normalized.isin(
                {
                    "FIXED INCOME",
                    "ALTERNATIVES",
                    "COMMODITIES",
                    "CASH & EQUIVALENTS",
                    "[CASH]",
                    "PREFERRED",
                    "CURRENCY FORWARDS",
                    "DERIVATIVES",
                }
            )
            support_keep |= indent.eq(6.0) & normalized.isin(
                {
                    "EQUITY - CANADIAN EQUITIES",
                    "EQUITY - INTERNATIONAL EQUITIES",
                    "EQUITY - US EQUITIES",
                }
            )
            support_keep |= comp_df["Breakdown Mapping Source"].isin(
                {
                    "Alternatives Detection",
                    "Other Detection",
                    "Cash Detection",
                }
            )
    elif section == "diversification":
        direct_mask = comp_df["Diversification Mapping Source"].eq("Direct Diversification Rule")
        fund_mask = comp_df["Diversification Mapping Source"].eq("Fund/Alternatives Detection")
        support_keep = (~has_hierarchy) | fund_mask | (direct_mask & indent.isin([4.0, 6.0, 8.0]))
    else:
        support_keep = pd.Series(True, index=comp_df.index)

    keep.loc[support_mask] = support_keep.loc[support_mask]
    return keep


def suppress_alternatives_wrappers(comp_df: pd.DataFrame) -> pd.DataFrame:
    if comp_df.empty or "Hierarchy Indent" not in comp_df.columns:
        return comp_df

    adjusted = comp_df.copy()

    for _, block in adjusted.groupby("Block Label", sort=False):
        rows = block.copy().reset_index()
        rows["Normalized Component"] = rows["Component"].map(normalize_key)
        indents = pd.to_numeric(rows["Hierarchy Indent"], errors="coerce")

        for row_idx, row in rows.iterrows():
            if row["Normalized Component"] != "ALTERNATIVES":
                continue

            current_indent = indents.iloc[row_idx]
            if pd.isna(current_indent):
                continue

            child_indices: List[int] = []
            lookahead = row_idx + 1
            while lookahead < len(rows):
                next_indent = indents.iloc[lookahead]
                if pd.notna(next_indent) and next_indent <= current_indent:
                    break
                child_indices.append(lookahead)
                lookahead += 1

            if not child_indices:
                continue

            child_rows = rows.loc[child_indices]
            child_weight_total = float(row["Port. Weight"])

            composition_child_total = child_rows.loc[
                child_rows["Composition Mapping Source"].eq("Funds/Alternatives Detection")
                & child_rows["Composition Group"].astype(str).str.strip().ne(""),
                "Port. Weight",
            ].sum()
            if abs(float(composition_child_total) - child_weight_total) <= 0.25:
                adjusted.at[row["index"], "Composition Group"] = ""
                adjusted.at[row["index"], "Composition Mapping Source"] = "Suppressed Alternatives Wrapper"

            breakdown_child_total = child_rows.loc[
                child_rows["Breakdown Mapping Source"].isin(
                    {
                        "Alternatives Detection",
                        "Other Detection",
                        "Cash Detection",
                    }
                )
                & child_rows["Breakdown Group"].astype(str).str.strip().ne(""),
                "Port. Weight",
            ].sum()
            if abs(float(breakdown_child_total) - child_weight_total) <= 0.25:
                adjusted.at[row["index"], "Breakdown Group"] = ""
                adjusted.at[row["index"], "Breakdown Mapping Source"] = "Suppressed Alternatives Wrapper"

    return adjusted


def parse_support_file(file_bytes: bytes, filename: str) -> Tuple[pd.DataFrame, Optional[pd.Timestamp]]:
    raw_df = read_tabular_sheet(file_bytes, filename, sheet_name=0, header=None)
    port_weight_pos = find_match_position(raw_df, "Port. Weight")
    non_composite_pos = find_match_position(raw_df, "Non-Composite")

    if port_weight_pos is None or non_composite_pos is None:
        raise ValueError(
            f"`{filename}` does not contain the expected `Port. Weight` / `Non-Composite` block."
        )

    start_row, name_col = non_composite_pos
    _, weight_col = port_weight_pos
    report_date = parse_support_report_date(raw_df, start_row)
    subset = raw_df.iloc[start_row:, [name_col, weight_col]].copy()
    subset.columns = ["Component", "Port. Weight"]
    subset["Sheet Row Number"] = subset.index + 1
    subset["Component"] = subset["Component"].apply(normalize_text)
    subset["Port. Weight"] = coerce_number_series(subset["Port. Weight"])
    subset = subset.dropna(how="all")
    subset = subset[subset["Component"] != ""]
    subset = subset[subset["Port. Weight"].notna()]
    component_key = subset["Component"].map(normalize_key)
    duplicate_adjacent = (
        component_key.eq(component_key.shift(-1))
        & subset["Port. Weight"].round(12).eq(subset["Port. Weight"].shift(-1).round(12))
    )
    subset = subset[~duplicate_adjacent]
    hierarchy_df = read_excel_hierarchy_metadata(file_bytes, filename, start_row, name_col)
    if not hierarchy_df.empty:
        subset = subset.merge(hierarchy_df, on="Sheet Row Number", how="left")
    else:
        subset["Hierarchy Indent"] = pd.NA
        subset["Is Summary Row"] = pd.NA
    subset = subset.drop(columns=["Sheet Row Number"])
    subset = subset.reset_index(drop=True)

    if subset.empty:
        raise ValueError(f"No support rows were parsed from `{filename}`.")

    return subset, report_date


def apply_composition_mapping(component: pd.Series) -> pd.Series:
    return classify_composition_mapping(component)["group"]


def classify_composition_mapping(component: pd.Series) -> pd.DataFrame:
    normalized = component.map(normalize_key)
    base = normalized.map(ASSET_CLASS_DETECTION)
    override = normalized.map(FUNDS_ALTERNATIVES_DETECTION)
    final = override.fillna(base).fillna("")
    source = pd.Series("", index=component.index, dtype="object")
    source.loc[base.notna()] = "Asset Class Detection"
    source.loc[override.notna()] = "Funds/Alternatives Detection"
    final = final.replace(
        {
            "Private Alt": "Private Alternatives",
            "Liquid Alt": "Private Alternatives",
        }
    )
    return pd.DataFrame({"group": final, "source": source}, index=component.index)


def apply_breakdown_mapping(component: pd.Series) -> pd.Series:
    return classify_breakdown_mapping(component)["group"]


def classify_breakdown_mapping(component: pd.Series) -> pd.DataFrame:
    normalized = component.map(normalize_key)
    result = pd.Series("", index=component.index, dtype="object")
    source = pd.Series("", index=component.index, dtype="object")
    result.loc[normalized.isin(BREAKDOWN_ALTERNATIVES)] = "Alternatives"
    source.loc[normalized.isin(BREAKDOWN_ALTERNATIVES)] = "Alternatives Detection"
    result.loc[normalized.isin(BREAKDOWN_CASH)] = "Cash"
    source.loc[normalized.isin(BREAKDOWN_CASH)] = "Cash Detection"
    result.loc[normalized.isin(BREAKDOWN_OTHER)] = "Other"
    source.loc[normalized.isin(BREAKDOWN_OTHER)] = "Other Detection"
    for key, label in BREAKDOWN_DIRECT.items():
        result.loc[normalized == key] = label
        source.loc[normalized == key] = "Direct Breakdown Rule"
    return pd.DataFrame({"group": result, "source": source}, index=component.index)


def get_diversification_fund_map() -> Dict[str, str]:
    fund_map: Dict[str, str] = {}
    non_alternative_overrides = {
        "INSTITUTIONAL FIDUCIARY TR MONEY MKT PTF": "Cash",
        "MSILF PRIME PORTFOLIO-INST": "Cash",
        "INVESCO PREMIER US GOV MONEY PTF": "Other",
    }

    # Diversification treats the underlying alternative sleeves themselves as
    # Alternatives, even when the higher-level breakdown logic classifies some
    # of those names as Other.
    for name in FUNDS_ALTERNATIVES_DETECTION:
        normalized = normalize_key(name)
        fund_map[name] = non_alternative_overrides.get(normalized, "Alternatives")

    for name in BREAKDOWN_ALTERNATIVES:
        fund_map[name] = "Alternatives"

    return fund_map


DIVERSIFICATION_FUND_MAP = get_diversification_fund_map()
DIVERSIFICATION_BREAKS = set(DIVERSIFICATION_DIRECT) | DIVERSIFICATION_WRAPPERS


def apply_diversification_mapping(comp_df: pd.DataFrame) -> pd.DataFrame:
    result = pd.Series("", index=comp_df.index, dtype="object")
    source = pd.Series("", index=comp_df.index, dtype="object")

    for _, block in comp_df.groupby("Block Label", sort=False):
        rows = block.copy().reset_index()
        rows["Normalized Component"] = rows["Component"].map(normalize_key)
        block_labels: List[str] = []
        block_sources: List[str] = []

        for _, row in rows.iterrows():
            normalized = row["Normalized Component"]
            if normalized in DIVERSIFICATION_DIRECT:
                block_labels.append(DIVERSIFICATION_DIRECT[normalized])
                block_sources.append("Direct Diversification Rule")
            elif normalized in DIVERSIFICATION_FUND_MAP:
                block_labels.append(DIVERSIFICATION_FUND_MAP[normalized])
                block_sources.append("Fund/Alternatives Detection")
            else:
                block_labels.append("")
                block_sources.append("")

        # Statement diversification treats TAA Alternatives at the wrapper level,
        # rather than expanding underlying private/alternative child funds.
        for row_idx, row in rows.iterrows():
            if row["Normalized Component"] != "ALTERNATIVES":
                continue

            is_taa_block = str(row.get("saa_taa", "")).upper() == "TAA"
            child_total = 0.0
            child_fund_indices: List[int] = []
            lookahead = row_idx + 1
            while lookahead < len(rows):
                normalized = rows.at[lookahead, "Normalized Component"]
                if normalized in DIVERSIFICATION_BREAKS:
                    break
                if normalized in DIVERSIFICATION_FUND_MAP:
                    child_total += float(rows.at[lookahead, "Port. Weight"])
                    child_fund_indices.append(lookahead)
                lookahead += 1

            if child_fund_indices and is_taa_block and abs(child_total - float(row["Port. Weight"])) <= 0.25:
                for child_idx in child_fund_indices:
                    block_labels[child_idx] = ""
                    block_sources[child_idx] = ""
            elif child_fund_indices and abs(child_total - float(row["Port. Weight"])) <= 0.25:
                block_labels[row_idx] = ""
                block_sources[row_idx] = ""

        result.loc[rows["index"]] = block_labels
        source.loc[rows["index"]] = block_sources

    return pd.DataFrame({"group": result, "source": source}, index=comp_df.index)


def build_diversification_summary(comp_df: pd.DataFrame, portfolio_total: float) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    warnings: List[str] = []
    diversification_rows = comp_df.copy()
    diversification_rows = diversification_rows[diversification_rows["Diversification Group"] != ""].copy()

    mapped_weight_by_block = diversification_rows.groupby("Block Label")["Port. Weight"].sum()
    residual_rows: List[dict] = []
    for block_label, block in comp_df.groupby("Block Label", sort=False):
        holding_mv = float(block["Total MV (CAD)"].iloc[0])
        residual_weight = 100.0 - float(mapped_weight_by_block.get(block_label, 0.0))
        if abs(residual_weight) <= 0.001:
            continue
        if residual_weight < -0.25:
            warnings.append(
                f"Diversification mapping over-allocated support block `{block_label}` by {abs(residual_weight):.2f}%."
            )
            continue
        residual_rows.append(
            {
                "Component": "Other Residual",
                "Port. Weight": residual_weight,
                "Fund Code": block["Fund Code"].iloc[0],
                "Fund Description": block["Fund Description"].iloc[0],
                "Total MV (CAD)": holding_mv,
                "mandate_code": block["mandate_code"].iloc[0],
                "support_code": block["support_code"].iloc[0],
                "saa_taa": block["saa_taa"].iloc[0],
                "Support File": block["Support File"].iloc[0],
                "Block Label": block_label,
                "Weighted MV (CAD)": holding_mv * residual_weight / 100.0,
                "Composition Group": "",
                "Breakdown Group": "",
                "Diversification Group": "Other",
                "Composition Mapping Source": "",
                "Breakdown Mapping Source": "",
                "Diversification Mapping Source": "Residual to Other",
                "Source Type": block["Source Type"].iloc[0],
            }
        )

    if residual_rows:
        diversification_rows = pd.concat(
            [diversification_rows, pd.DataFrame(residual_rows)],
            ignore_index=True,
        )

    sector_totals = (
        diversification_rows[
            diversification_rows["Diversification Group"].isin(DIVERSIFICATION_SECTOR_GROUPS)
        ]
        .groupby("Diversification Group")["Weighted MV (CAD)"]
        .sum()
    )
    low_sector_groups = set(
        sector_totals[(sector_totals / portfolio_total * 100.0) < 2.0].index.tolist()
    )
    if low_sector_groups:
        low_sector_mask = diversification_rows["Diversification Group"].isin(low_sector_groups)
        diversification_rows.loc[low_sector_mask, "Diversification Group"] = "Other"
        diversification_rows.loc[low_sector_mask, "Diversification Mapping Source"] = (
            "Sector under 2% to Other"
        )

    diversification_pivot = (
        diversification_rows.pivot_table(
            index="Diversification Group",
            columns="saa_taa",
            values="Weighted MV (CAD)",
            aggfunc="sum",
            fill_value=0.0,
        )
        .reindex(DIVERSIFICATION_GROUP_ORDER, fill_value=0.0)
    )
    for column in ["SAA", "TAA"]:
        if column not in diversification_pivot.columns:
            diversification_pivot[column] = 0.0
    diversification_pivot = diversification_pivot[["SAA", "TAA"]]

    diversification_df = pd.DataFrame(
        {
            "Portfolio Diversification": DIVERSIFICATION_GROUP_ORDER,
            "Strategic Asset Allocation %": diversification_pivot["SAA"].values / portfolio_total * 100.0,
            "Tactical Asset Allocation %": diversification_pivot["TAA"].values / portfolio_total * 100.0,
        }
    )
    diversification_df["Portfolio %"] = (
        diversification_df["Strategic Asset Allocation %"] + diversification_df["Tactical Asset Allocation %"]
    )
    diversification_df = diversification_df[diversification_df["Portfolio %"].abs() > 0.000001].copy()
    diversification_total_row = pd.DataFrame(
        [
            {
                "Portfolio Diversification": "Total Portfolio Diversification",
                "Strategic Asset Allocation %": diversification_df["Strategic Asset Allocation %"].sum(),
                "Tactical Asset Allocation %": diversification_df["Tactical Asset Allocation %"].sum(),
                "Portfolio %": diversification_df["Portfolio %"].sum(),
            }
        ]
    )
    diversification_df = pd.concat([diversification_df, diversification_total_row], ignore_index=True)
    diversification_df = sort_report_rows_by_percentage(
        diversification_df,
        "Portfolio %",
        "Portfolio Diversification",
        "Total Portfolio Diversification",
    )

    return diversification_rows, diversification_df, warnings


def build_uploaded_support_map(files: Iterable[object]) -> Tuple[Dict[str, dict], List[str]]:
    support_map: Dict[str, dict] = {}
    warnings: List[str] = []

    for uploaded in files:
        code = extract_mandate_code(uploaded.name)
        if not code:
            warnings.append(f"Ignored support file `{uploaded.name}` because no 5-digit mandate code was found in the filename.")
            continue
        if code in support_map:
            warnings.append(f"Duplicate support file uploaded for mandate code {code}. Using `{support_map[code]['filename']}` and ignoring `{uploaded.name}`.")
            continue
        support_map[code] = {
            "filename": uploaded.name,
            "bytes": uploaded.getvalue(),
        }

    return support_map, warnings


def add_support_file_record(
    support_map: Dict[str, dict],
    warnings: List[str],
    filename: str,
    file_bytes: bytes,
) -> None:
    code = extract_mandate_code(filename)
    if not code:
        warnings.append(f"Ignored support file `{filename}` because no 5-digit code was found in the filename.")
        return
    if code in support_map:
        warnings.append(
            f"Duplicate support file saved for code {code}. Using `{support_map[code]['filename']}` and ignoring `{filename}`."
        )
        return
    support_map[code] = {
        "filename": filename,
        "bytes": file_bytes,
    }


def build_saved_support_map(saved_files: List[dict]) -> Tuple[Dict[str, dict], List[str]]:
    support_map: Dict[str, dict] = {}
    warnings: List[str] = []

    for saved in saved_files:
        filename = saved["filename"]
        if filename.lower().endswith(".zip"):
            try:
                with ZipFile(BytesIO(saved["bytes"])) as archive:
                    for member in archive.namelist():
                        if member.endswith("/") or not member.lower().endswith((".xlsx", ".xls", ".xlsm", ".csv")):
                            continue
                        add_support_file_record(
                            support_map,
                            warnings,
                            f"{filename} :: {member}",
                            archive.read(member),
                        )
            except Exception as exc:
                warnings.append(f"Ignored ZIP file `{filename}` because it could not be read: {exc}")
            continue

        add_support_file_record(support_map, warnings, filename, saved["bytes"])

    return support_map, warnings


def resolve_support_file(
    holding: pd.Series,
    support_map: Dict[str, dict],
    factset_override: Optional[dict] = None,
) -> Tuple[Optional[str], Optional[dict], List[str]]:
    candidates = build_support_candidates(
        holding["Fund Code"],
        holding.get("mandate_code"),
        holding.get("saa_taa"),
        fund_description=holding.get("Fund Description"),
        factset_override=factset_override,
    )
    for code in candidates:
        if code in support_map:
            return code, support_map[code], candidates
    return None, None, candidates


def calculate_reports(
    holdings: pd.DataFrame,
    support_map: Dict[str, dict],
    sma_override: Optional[dict] = None,
    factset_override: Optional[dict] = None,
) -> Tuple[dict, List[str], List[str]]:
    warnings: List[str] = []
    info: List[str] = []
    blocks: List[pd.DataFrame] = []
    matched_codes: set[str] = set()
    unresolved_holdings: List[str] = []
    reporting_periods: set[str] = set()

    for _, holding in holdings.iterrows():
        code, support_file, candidates = resolve_support_file(
            holding,
            support_map,
            factset_override=factset_override,
        )
        if support_file is None or code is None:
            sma_row = get_sma_row_for_holding(holding, sma_override=sma_override)
            if sma_row is not None:
                blocks.append(build_sma_rows(holding, sma_row))
                continue

            descriptor = f"{holding['Fund Description']} (expected one of: {', '.join(candidates) if candidates else 'no derived support code'})"
            unresolved_holdings.append(descriptor)
            continue

        matched_codes.add(code)
        support_rows, report_date = parse_support_file(support_file["bytes"], support_file["filename"])
        reporting_period = format_reporting_period(report_date)
        if reporting_period:
            reporting_periods.add(reporting_period)
        support_rows["Fund Code"] = holding["Fund Code"]
        support_rows["Fund Description"] = holding["Fund Description"]
        support_rows["Total MV (CAD)"] = float(holding["Total MV (CAD)"])
        support_rows["mandate_code"] = holding.get("mandate_code")
        support_rows["support_code"] = code
        support_rows["saa_taa"] = holding["allocation_bucket"]
        support_rows["Holding Type"] = holding["saa_taa"]
        support_rows["Support File"] = support_file["filename"]
        support_rows["Block Label"] = f"{holding['Fund Description']} - {code} - {holding['saa_taa']}"
        support_rows["Weighted MV (CAD)"] = support_rows["Total MV (CAD)"] * support_rows["Port. Weight"] / 100.0
        support_rows["Source Type"] = "Support File"
        blocks.append(support_rows)

    if unresolved_holdings:
        raise ValueError(
            "Missing support files for holdings: " + "; ".join(unresolved_holdings)
        )

    extra_codes = sorted(set(support_map) - matched_codes)
    if extra_codes:
        ignored = [support_map[code]["filename"] for code in extra_codes]
        info.append("Ignored extra uploaded support files not referenced in holdings: " + ", ".join(ignored))

    comp_df = pd.concat(blocks, ignore_index=True)
    if "Composition Group" not in comp_df.columns:
        comp_df["Composition Group"] = ""
    if "Composition Mapping Source" not in comp_df.columns:
        comp_df["Composition Mapping Source"] = ""
    if "Breakdown Group" not in comp_df.columns:
        comp_df["Breakdown Group"] = ""
    if "Breakdown Mapping Source" not in comp_df.columns:
        comp_df["Breakdown Mapping Source"] = ""
    if "Diversification Group" not in comp_df.columns:
        comp_df["Diversification Group"] = ""
    if "Diversification Mapping Source" not in comp_df.columns:
        comp_df["Diversification Mapping Source"] = ""
    if "Hierarchy Indent" not in comp_df.columns:
        comp_df["Hierarchy Indent"] = pd.NA
    if "Is Summary Row" not in comp_df.columns:
        comp_df["Is Summary Row"] = pd.NA
    for column in [
        "Composition Group",
        "Composition Mapping Source",
        "Breakdown Group",
        "Breakdown Mapping Source",
        "Diversification Group",
        "Diversification Mapping Source",
    ]:
        comp_df[column] = comp_df[column].fillna("")
    composition_blank = comp_df["Composition Group"].astype(str).str.strip().eq("")
    breakdown_blank = comp_df["Breakdown Group"].astype(str).str.strip().eq("")
    composition_classification = classify_composition_mapping(comp_df.loc[composition_blank, "Component"])
    comp_df.loc[composition_blank, "Composition Group"] = composition_classification["group"]
    comp_df.loc[composition_blank, "Composition Mapping Source"] = composition_classification["source"]
    breakdown_classification = classify_breakdown_mapping(comp_df.loc[breakdown_blank, "Component"])
    comp_df.loc[breakdown_blank, "Breakdown Group"] = breakdown_classification["group"]
    comp_df.loc[breakdown_blank, "Breakdown Mapping Source"] = breakdown_classification["source"]
    comp_df = suppress_alternatives_wrappers(comp_df)
    diversification_classification = apply_diversification_mapping(comp_df)
    comp_df["Diversification Group"] = diversification_classification["group"]
    comp_df["Diversification Mapping Source"] = diversification_classification["source"]

    portfolio_total = float(holdings["Total MV (CAD)"].sum())
    if portfolio_total <= 0:
        raise ValueError("The total holdings market value must be greater than 0.")

    composition_rows = comp_df[
        comp_df["Composition Group"].ne("")
        & select_support_rows_by_hierarchy(comp_df, "composition")
    ].copy()
    breakdown_rows = comp_df[
        comp_df["Breakdown Group"].ne("")
        & select_support_rows_by_hierarchy(comp_df, "breakdown")
    ].copy()
    composition_pivot = (
        composition_rows.assign(
            **{
                "Composition Source": composition_rows["Source Type"].map(
                    {"Support File": "Actively Managed", "SMA": "Separately Managed"}
                ).fillna("Actively Managed")
            }
        )
        .pivot_table(
            index="Composition Group",
            columns="Composition Source",
            values="Weighted MV (CAD)",
            aggfunc="sum",
            fill_value=0.0,
        )
        .reindex(COMPOSITION_GROUP_ORDER, fill_value=0.0)
    )
    for column in ["Actively Managed", "Separately Managed"]:
        if column not in composition_pivot.columns:
            composition_pivot[column] = 0.0
    composition_pivot = composition_pivot[["Actively Managed", "Separately Managed"]]
    composition_summary = composition_pivot.sum(axis=1)
    composition_df = pd.DataFrame(
        {
            "Asset Classes": COMPOSITION_GROUP_ORDER,
            "Portfolio Market Value (CDN)": composition_summary.values,
            "Actively Managed Market Value (CDN)": composition_pivot["Actively Managed"].values,
            "Separately Managed Market Value (CDN)": composition_pivot["Separately Managed"].values,
            "% of Portfolio (CDN)": composition_summary.values / portfolio_total * 100.0,
        }
    )
    composition_df = composition_df[composition_df["Portfolio Market Value (CDN)"].abs() > 0.0001].reset_index(drop=True)
    composition_total_row = pd.DataFrame(
        [
            {
                "Asset Classes": "Total Market Value of Asset Classes",
                "Portfolio Market Value (CDN)": composition_df["Portfolio Market Value (CDN)"].sum(),
                "Actively Managed Market Value (CDN)": composition_df["Actively Managed Market Value (CDN)"].sum(),
                "Separately Managed Market Value (CDN)": composition_df["Separately Managed Market Value (CDN)"].sum(),
                "% of Portfolio (CDN)": composition_df["% of Portfolio (CDN)"].sum(),
            }
        ]
    )
    composition_df = pd.concat([composition_df, composition_total_row], ignore_index=True)
    composition_df = sort_report_rows_by_percentage(
        composition_df,
        "% of Portfolio (CDN)",
        "Asset Classes",
        "Total Market Value of Asset Classes",
    )

    actively_managed_total = float(composition_pivot["Actively Managed"].sum())
    breakdown_denominator = actively_managed_total if actively_managed_total > 0 else portfolio_total
    breakdown_rows = breakdown_rows[breakdown_rows["Source Type"].eq("Support File")].copy()
    diversification_source_rows = comp_df[
        comp_df["Source Type"].eq("Support File")
        & select_support_rows_by_hierarchy(comp_df, "diversification")
    ].copy()
    diversification_rows, diversification_df, diversification_warnings = build_diversification_summary(
        diversification_source_rows,
        breakdown_denominator,
    )
    warnings.extend(diversification_warnings)

    breakdown_pivot = (
        breakdown_rows.pivot_table(
            index="Breakdown Group",
            columns="saa_taa",
            values="Weighted MV (CAD)",
            aggfunc="sum",
            fill_value=0.0,
        )
        .reindex(BREAKDOWN_GROUP_ORDER, fill_value=0.0)
    )
    for column in ["SAA", "TAA"]:
        if column not in breakdown_pivot.columns:
            breakdown_pivot[column] = 0.0
    breakdown_pivot = breakdown_pivot[["SAA", "TAA"]]

    breakdown_df = pd.DataFrame(
        {
            "Actively Managed Asset Classes": BREAKDOWN_GROUP_ORDER,
            "Strategic Asset Allocation %": breakdown_pivot["SAA"].values / breakdown_denominator * 100.0,
            "Tactical Asset Allocation %": breakdown_pivot["TAA"].values / breakdown_denominator * 100.0,
        }
    )
    breakdown_df["Portfolio %"] = (
        breakdown_df["Strategic Asset Allocation %"] + breakdown_df["Tactical Asset Allocation %"]
    )
    breakdown_total_row = pd.DataFrame(
        [
            {
                "Actively Managed Asset Classes": "Total of Actively Managed Assets",
                "Strategic Asset Allocation %": breakdown_df["Strategic Asset Allocation %"].sum(),
                "Tactical Asset Allocation %": breakdown_df["Tactical Asset Allocation %"].sum(),
                "Portfolio %": breakdown_df["Portfolio %"].sum(),
            }
        ]
    )
    breakdown_df = pd.concat([breakdown_df, breakdown_total_row], ignore_index=True)
    breakdown_df = sort_report_rows_by_percentage(
        breakdown_df,
        "Portfolio %",
        "Actively Managed Asset Classes",
        "Total of Actively Managed Assets",
    )

    composition_mapped_total = composition_rows["Weighted MV (CAD)"].sum()
    if abs(composition_mapped_total - portfolio_total) > 0.5:
        unmatched = portfolio_total - composition_mapped_total
        warnings.append(
            f"Composition mapping did not reconcile exactly to holdings total. Unmapped amount: ${unmatched:,.2f}."
        )

    breakdown_mapped_total = breakdown_rows["Weighted MV (CAD)"].sum()
    if abs(breakdown_mapped_total - breakdown_denominator) > 0.5:
        unmatched = breakdown_denominator - breakdown_mapped_total
        warnings.append(
            f"Breakdown mapping did not reconcile exactly to actively managed total. Unmapped amount: ${unmatched:,.2f}."
        )

    comp_block_check = composition_rows.groupby("Block Label")["Port. Weight"].sum()
    off_comp = comp_block_check[comp_block_check.sub(100).abs() > 0.25]
    if not off_comp.empty:
        warnings.append(
            "Some support blocks do not sum to 100% after composition mapping: "
            + ", ".join(f"{label} ({value:.2f}%)" for label, value in off_comp.items())
        )

    break_block_check = breakdown_rows.groupby("Block Label")["Port. Weight"].sum()
    off_break = break_block_check[break_block_check.sub(100).abs() > 0.25]
    if not off_break.empty:
        warnings.append(
            "Some support blocks do not sum to 100% after breakdown mapping: "
            + ", ".join(f"{label} ({value:.2f}%)" for label, value in off_break.items())
        )

    equity_composition = float(composition_pivot.loc["Equity", "Actively Managed"])
    equity_breakdown = float(
        breakdown_pivot.loc[["International Equity", "US Equity", "Canadian Equity"]].sum().sum()
    )
    if abs(equity_composition - equity_breakdown) > 0.5:
        warnings.append(
            f"Equity reconciliation failed. Composition Equity = ${equity_composition:,.2f}, Breakdown Equity Total = ${equity_breakdown:,.2f}."
        )

    diversification_mapped_total = diversification_rows["Weighted MV (CAD)"].sum()
    if abs(diversification_mapped_total - breakdown_denominator) > 0.5:
        unmatched = breakdown_denominator - diversification_mapped_total
        warnings.append(
            f"Diversification mapping did not reconcile exactly to actively managed total. Unmapped amount: ${unmatched:,.2f}."
        )

    diversification_block_check = diversification_rows.groupby("Block Label")["Port. Weight"].sum()
    off_diversification = diversification_block_check[diversification_block_check.sub(100).abs() > 0.25]
    if not off_diversification.empty:
        warnings.append(
            "Some support blocks do not sum to 100% after diversification mapping: "
            + ", ".join(f"{label} ({value:.2f}%)" for label, value in off_diversification.items())
        )

    if holdings["allocation_bucket"].eq("SAA").all():
        info.append("This is an SAA-only portfolio. Tactical values remain at 0.00%.")
    if holdings["allocation_bucket"].eq("TAA").all():
        info.append("This is a TAA-only portfolio. Strategic values remain at 0.00%.")

    factset_models, factset_warnings, factset_info = detect_factset_models(
        holdings,
        factset_override=factset_override,
    )
    warnings.extend(factset_warnings)
    info.extend(factset_info)

    reporting_period = None
    if len(reporting_periods) == 1:
        reporting_period = next(iter(reporting_periods))
        info.append(f"Detected support-file reporting period: {reporting_period}.")
    elif len(reporting_periods) > 1:
        warnings.append(
            "Uploaded support files appear to span multiple reporting periods: "
            + ", ".join(sorted(reporting_periods))
            + "."
        )
    elif not comp_df.empty and (comp_df["Source Type"] == "Support File").any():
        warnings.append(
            "No reporting period could be detected from the uploaded support files."
        )

    return {
        "holdings": holdings.copy(),
        "comp_detail": comp_df,
        "composition": composition_df,
        "breakdown": breakdown_df,
        "diversification_detail": diversification_rows,
        "diversification": diversification_df,
        "portfolio_total": portfolio_total,
        "reporting_period": reporting_period,
        "factset_models": factset_models,
    }, warnings, info


def format_currency(value: float) -> str:
    return f"${value:,.2f}"


def format_percent(value: float) -> str:
    return f"{value:,.2f}%"


def sort_report_rows_by_percentage(
    df: pd.DataFrame,
    percent_column: str,
    total_label_column: str,
    total_label: str,
) -> pd.DataFrame:
    if df.empty or percent_column not in df.columns:
        return df
    total_rows = df[df[total_label_column].eq(total_label)]
    detail_rows = df[~df[total_label_column].eq(total_label)].copy()
    detail_rows = detail_rows.sort_values(percent_column, ascending=False, kind="mergesort")
    return pd.concat([detail_rows, total_rows], ignore_index=True)


def build_composition_display(df: pd.DataFrame) -> pd.DataFrame:
    display = df.copy()
    for col in [
        "Portfolio Market Value (CDN)",
        "Actively Managed Market Value (CDN)",
        "Separately Managed Market Value (CDN)",
    ]:
        if col in display.columns:
            display[col] = display[col].map(format_currency)
    display["% of Portfolio (CDN)"] = display["% of Portfolio (CDN)"].map(format_percent)
    return display


def build_breakdown_display(df: pd.DataFrame) -> pd.DataFrame:
    display = df.copy()
    for col in ["Strategic Asset Allocation %", "Tactical Asset Allocation %", "Portfolio %"]:
        display[col] = display[col].map(format_percent)
    return display


def build_diversification_display(df: pd.DataFrame) -> pd.DataFrame:
    display = df.copy()
    for col in ["Strategic Asset Allocation %", "Tactical Asset Allocation %", "Portfolio %"]:
        display[col] = display[col].map(format_percent)
    return display


def build_mapping_audit_display(df: pd.DataFrame) -> pd.DataFrame:
    display = df.copy()
    if "Max Port. Weight" in display.columns:
        display["Max Port. Weight"] = display["Max Port. Weight"].map(format_percent)
    return display


def build_other_review_display(df: pd.DataFrame) -> pd.DataFrame:
    display = df.copy()
    if "Port. Weight" in display.columns:
        display["Port. Weight"] = display["Port. Weight"].map(format_percent)
    if "Weighted MV (CAD)" in display.columns:
        display["Weighted MV (CAD)"] = display["Weighted MV (CAD)"].map(format_currency)
    return display


def build_unreviewed_labels_display(df: pd.DataFrame) -> pd.DataFrame:
    display = df.copy()
    if "Max Port. Weight" in display.columns:
        display["Max Port. Weight"] = display["Max Port. Weight"].map(format_percent)
    return display


def build_composition_chart(df: pd.DataFrame) -> go.Figure:
    chart_df = df[df["Asset Classes"] != "Total Market Value of Asset Classes"].copy()
    chart_df = chart_df[chart_df["Portfolio Market Value (CDN)"] > 0].copy()
    fig = px.pie(
        chart_df,
        names="Asset Classes",
        values="Portfolio Market Value (CDN)",
        hole=0.6,
        color="Asset Classes",
        color_discrete_map={
            "Income": PRIMARY_BLUE,
            "Equity": "#4B4B4B",
            "Balanced": "#7B6F8E",
            "Liquid Alternatives": "#6E8B87",
            "Sector": "#9A8F6B",
            "Cash": "#D7D1C3",
            "Other": "#A7A7A7",
            "Private Alternatives": "#C7B168",
        },
    )
    fig.update_traces(
        textposition="outside",
        texttemplate="%{label} %{percent:.2%}",
        hovertemplate="<b>%{label}</b><br>Market Value: $%{value:,.2f}<br>Portfolio: %{percent:.2%}<extra></extra>",
        textfont=dict(color="#111827", size=16),
        marker=dict(line=dict(color="white", width=2)),
        sort=False,
        direction="clockwise",
    )
    fig.update_layout(
        margin=dict(l=90, r=120, t=30, b=40),
        showlegend=False,
        paper_bgcolor="white",
        plot_bgcolor="white",
        font=dict(color="#111827", size=16),
        uniformtext=dict(minsize=14, mode="hide"),
    )
    fig.update_traces(automargin=True)
    return fig


def build_stacked_allocation_chart(
    df: pd.DataFrame,
    category_column: str,
    total_label: str,
    left_margin: int = 180,
    right_margin: int = 120,
    top_margin: int = 70,
    internal_label_threshold: float = 4.0,
) -> go.Figure:
    chart_df = df[df[category_column] != total_label].copy()
    long_df = chart_df.melt(
        id_vars=category_column,
        value_vars=["Strategic Asset Allocation %", "Tactical Asset Allocation %"],
        var_name="Allocation Type",
        value_name="Percent",
    )
    long_df["Display Percent"] = long_df["Percent"].apply(
        lambda value: value if float(value) >= internal_label_threshold else None
    )
    fig = px.bar(
        long_df,
        x="Percent",
        y=category_column,
        color="Allocation Type",
        orientation="h",
        barmode="relative",
        text="Display Percent",
        color_discrete_map={
            "Strategic Asset Allocation %": PRIMARY_BLUE,
            "Tactical Asset Allocation %": SECONDARY_GRAY,
        },
        category_orders={
            category_column: chart_df[category_column].tolist(),
        },
    )
    fig.update_traces(
        texttemplate="%{text:.2f}%",
        textposition="inside",
        insidetextanchor="middle",
        hovertemplate="<b>%{y}</b><br>%{fullData.name}: %{x:.2f}%<extra></extra>",
    )
    fig.update_traces(
        textfont=dict(color="white", size=17),
        selector=dict(name="Strategic Asset Allocation %"),
    )
    fig.update_traces(
        textfont=dict(color="#1F2937", size=17),
        selector=dict(name="Tactical Asset Allocation %"),
    )
    totals = chart_df["Portfolio %"].tolist()
    labels = chart_df[category_column].tolist()
    max_total = max(totals) if totals else 0.0
    x_axis_max = max(max_total * 1.24, max_total + 3.5, 10.0)
    fig.update_layout(
        margin=dict(l=left_margin, r=right_margin, t=top_margin, b=40),
        legend=dict(
            title="",
            orientation="h",
            yanchor="bottom",
            y=1.04,
            xanchor="left",
            x=0,
            font=dict(color="#111827", size=16),
        ),
        xaxis_title="Portfolio %",
        yaxis_title="",
        paper_bgcolor="white",
        plot_bgcolor="white",
        font=dict(color="#111827", size=16),
        xaxis=dict(
            range=[0, x_axis_max],
            automargin=True,
            tickfont=dict(color="#111827", size=16),
            title_font=dict(color="#111827", size=16),
            gridcolor="#E5E7EB",
            zerolinecolor="#D1D5DB",
        ),
        yaxis=dict(
            automargin=True,
            tickfont=dict(color="#111827", size=18),
            title_font=dict(color="#111827", size=16),
        ),
        uniformtext=dict(minsize=12, mode="hide"),
    )
    fig.add_trace(
        go.Scatter(
            x=totals,
            y=labels,
            mode="text",
            text=[f"{value:.2f}%" for value in totals],
            textposition="middle right",
            textfont=dict(color="#111827", size=18),
            showlegend=False,
            hoverinfo="skip",
            cliponaxis=False,
        )
    )
    return fig


def build_breakdown_chart(df: pd.DataFrame) -> go.Figure:
    return build_stacked_allocation_chart(
        df,
        category_column="Actively Managed Asset Classes",
        total_label="Total of Actively Managed Assets",
        left_margin=170,
        right_margin=150,
        top_margin=80,
        internal_label_threshold=4.0,
    )


def build_diversification_chart(df: pd.DataFrame) -> go.Figure:
    chart_df = df[df["Portfolio Diversification"] != "Total Portfolio Diversification"].copy()
    strategic_total = float(chart_df["Strategic Asset Allocation %"].sum())
    tactical_total = float(chart_df["Tactical Asset Allocation %"].sum())

    strategic_labels = [
        f"{value:.2f}%" if value >= 1.0 else ""
        for value in chart_df["Strategic Asset Allocation %"].tolist()
    ]
    tactical_labels = [
        f"{value:.2f}%" if value >= 1.0 else ""
        for value in chart_df["Tactical Asset Allocation %"].tolist()
    ]
    totals = chart_df["Portfolio %"].tolist()
    categories = chart_df["Portfolio Diversification"].tolist()
    max_total = max(totals) if totals else 0.0
    x_axis_max = max(max_total * 1.14, 10.0)

    fig = go.Figure()
    fig.add_trace(
        go.Bar(
            x=chart_df["Strategic Asset Allocation %"],
            y=categories,
            orientation="h",
            name=f"Strategic Asset Allocation {strategic_total:.2f}%",
            marker_color=PRIMARY_BLUE,
            text=strategic_labels,
            textposition="inside",
            insidetextanchor="middle",
            textfont=dict(color="white", size=17),
            hovertemplate="<b>%{y}</b><br>%{fullData.name}: %{x:.2f}%<extra></extra>",
        )
    )
    fig.add_trace(
        go.Bar(
            x=chart_df["Tactical Asset Allocation %"],
            y=categories,
            orientation="h",
            name=f"Tactical Asset Allocation {tactical_total:.2f}%",
            marker_color=SECONDARY_GRAY,
            text=tactical_labels,
            textposition="inside",
            insidetextanchor="middle",
            textfont=dict(color="#374151", size=17),
            hovertemplate="<b>%{y}</b><br>%{fullData.name}: %{x:.2f}%<extra></extra>",
        )
    )
    fig.add_trace(
        go.Scatter(
            x=totals,
            y=categories,
            mode="text",
            text=[f"{value:.2f}%" for value in totals],
            textposition="middle right",
            textfont=dict(color="#111827", size=19),
            showlegend=False,
            hoverinfo="skip",
            cliponaxis=False,
        )
    )
    fig.update_layout(
        barmode="relative",
        margin=dict(l=250, r=120, t=70, b=50),
        legend=dict(
            title="",
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="left",
            x=0,
            font=dict(color="#111827", size=16),
        ),
        xaxis_title="Portfolio %",
        yaxis_title="",
        paper_bgcolor="white",
        plot_bgcolor="white",
        font=dict(color="#111827", size=16),
        xaxis=dict(
            range=[0, x_axis_max],
            automargin=True,
            tickfont=dict(color="#111827", size=15),
            title_font=dict(color="#111827", size=16),
            gridcolor="#E5E7EB",
            zeroline=False,
        ),
        yaxis=dict(
            automargin=True,
            tickfont=dict(color="#111827", size=17),
            categoryorder="array",
            categoryarray=list(reversed(categories)),
        ),
        uniformtext=dict(minsize=12, mode="hide"),
    )
    return fig


def build_audit_tables(results: dict) -> dict:
    comp_detail = results["comp_detail"].copy()
    diversification_detail = results["diversification_detail"].copy()

    mapping_rows = comp_detail[
        (comp_detail["Composition Mapping Source"] != "")
        | (comp_detail["Breakdown Mapping Source"] != "")
        | (comp_detail["Diversification Mapping Source"] != "")
    ].copy()
    mapping_audit = (
        mapping_rows.groupby("Component", as_index=False)
        .agg(
            {
                "Support File": lambda values: ", ".join(sorted({str(value) for value in values if normalize_text(value)})),
                "Port. Weight": "max",
                "Composition Group": lambda values: ", ".join(sorted({str(value) for value in values if normalize_text(value)})),
                "Composition Mapping Source": lambda values: ", ".join(sorted({str(value) for value in values if normalize_text(value)})),
                "Breakdown Group": lambda values: ", ".join(sorted({str(value) for value in values if normalize_text(value)})),
                "Breakdown Mapping Source": lambda values: ", ".join(sorted({str(value) for value in values if normalize_text(value)})),
                "Diversification Group": lambda values: ", ".join(sorted({str(value) for value in values if normalize_text(value)})),
                "Diversification Mapping Source": lambda values: ", ".join(sorted({str(value) for value in values if normalize_text(value)})),
            }
        )
        .sort_values(["Port. Weight", "Component"], ascending=[False, True])
        .reset_index(drop=True)
    )
    mapping_audit = mapping_audit.rename(columns={"Port. Weight": "Max Port. Weight"})

    other_review = diversification_detail[
        (diversification_detail["Diversification Group"] == "Other")
        | (diversification_detail["Diversification Mapping Source"] == "Residual to Other")
    ].copy()
    other_review = other_review[
        [
            "Block Label",
            "Component",
            "Port. Weight",
            "Weighted MV (CAD)",
            "saa_taa",
            "Diversification Group",
            "Diversification Mapping Source",
        ]
    ].sort_values(["Block Label", "Port. Weight"], ascending=[True, False]).reset_index(drop=True)

    large_unreviewed = comp_detail[
        (comp_detail["Composition Mapping Source"] == "")
        & (comp_detail["Breakdown Mapping Source"] == "")
        & (comp_detail["Diversification Mapping Source"] == "")
        & (comp_detail["Source Type"] == "Support File")
        & (comp_detail["Port. Weight"].abs() >= 0.25)
    ].copy()
    unreviewed_labels = (
        large_unreviewed.groupby("Component", as_index=False)
        .agg(
            {
                "Support File": lambda values: ", ".join(sorted({str(value) for value in values if normalize_text(value)})),
                "Port. Weight": "max",
            }
        )
        .rename(columns={"Port. Weight": "Max Port. Weight"})
        .sort_values(["Max Port. Weight", "Component"], ascending=[False, True])
        .reset_index(drop=True)
    )

    summary = {
        "Mapped Labels": int(mapping_audit["Component"].nunique()) if not mapping_audit.empty else 0,
        "Other Review Rows": len(other_review),
        "Large Unreviewed Labels": int(unreviewed_labels["Component"].nunique()) if not unreviewed_labels.empty else 0,
    }

    return {
        "summary": summary,
        "mapping_audit": mapping_audit,
        "other_review": other_review,
        "unreviewed_labels": unreviewed_labels,
    }


def maybe_render_figure_png(fig: go.Figure, width: int, height: int) -> Optional[bytes]:
    try:
        return fig.to_image(format="png", width=width, height=height, scale=2)
    except Exception:
        return None


def autofit_columns(worksheet) -> None:
    from openpyxl.utils import get_column_letter

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)


def write_table(
    worksheet,
    start_row: int,
    dataframe: pd.DataFrame,
    currency_columns: Optional[Iterable[str]] = None,
    percent_columns: Optional[Iterable[str]] = None,
    title: Optional[str] = None,
    highlight_total: bool = True,
) -> int:
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_font = Font(bold=True)
    row = start_row
    if title:
        worksheet.cell(row=row, column=1, value=title).font = Font(size=14, bold=True)
        row += 2

    for col_idx, column_name in enumerate(dataframe.columns, start=1):
        cell = worksheet.cell(row=row, column=col_idx, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row += 1
    currency_columns = set(currency_columns or [])
    percent_columns = set(percent_columns or [])

    for record_idx, (_, record) in enumerate(dataframe.iterrows(), start=0):
        is_total = highlight_total and record_idx == len(dataframe) - 1
        for col_idx, column_name in enumerate(dataframe.columns, start=1):
            cell = worksheet.cell(row=row, column=col_idx, value=record[column_name])
            if column_name in currency_columns:
                cell.number_format = '$#,##0.00'
            elif column_name in percent_columns:
                cell.number_format = '0.00%'
            if is_total:
                cell.fill = total_fill
                cell.font = total_font
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
        row += 1

    autofit_columns(worksheet)
    worksheet.freeze_panes = worksheet["A3"] if title else worksheet["A2"]
    return row


def write_audit_sheet(
    worksheet,
    title: str,
    dataframe: pd.DataFrame,
    currency_columns: Optional[Iterable[str]] = None,
    percent_columns: Optional[Iterable[str]] = None,
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    worksheet.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)
    worksheet.cell(row=2, column=1, value="Formula cells are intentionally included so the calculation can be traced row by row.").font = Font(italic=True)
    write_table(
        worksheet,
        start_row=4,
        dataframe=dataframe,
        currency_columns=currency_columns,
        percent_columns=percent_columns,
        highlight_total=False,
    )
    header_fill = PatternFill(fill_type="solid", fgColor="DDEBFF")
    for row in worksheet.iter_rows(min_row=4, max_row=4):
        for cell in row:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.freeze_panes = worksheet["A5"]


def build_summary_audit(
    holdings_df: pd.DataFrame,
    composition_df: pd.DataFrame,
    breakdown_df: pd.DataFrame,
    diversification_df: pd.DataFrame,
) -> pd.DataFrame:
    portfolio_total = float(holdings_df["Total MV (CAD)"].sum())
    composition_total = composition_df.iloc[-1]
    actively_managed_total = float(composition_total.get("Actively Managed Market Value (CDN)", portfolio_total))
    separately_managed_total = float(composition_total.get("Separately Managed Market Value (CDN)", 0.0))

    records: List[dict] = [
        {
            "Section": "Denominator",
            "Line Item": "Full portfolio market value",
            "SAA MV": "",
            "TAA MV": "",
            "Total MV": portfolio_total,
            "Denominator": portfolio_total,
            "SAA %": "",
            "TAA %": "",
            "Portfolio %": 1.0,
            "Explanation": "Used for Portfolio Composition.",
        },
        {
            "Section": "Denominator",
            "Line Item": "Actively managed market value",
            "SAA MV": "",
            "TAA MV": "",
            "Total MV": actively_managed_total,
            "Denominator": actively_managed_total,
            "SAA %": "",
            "TAA %": "",
            "Portfolio %": 1.0,
            "Explanation": "Used for Portfolio Breakdown and Portfolio Diversification.",
        },
        {
            "Section": "Denominator",
            "Line Item": "Separately managed market value",
            "SAA MV": "",
            "TAA MV": "",
            "Total MV": separately_managed_total,
            "Denominator": portfolio_total,
            "SAA %": "",
            "TAA %": "",
            "Portfolio %": separately_managed_total / portfolio_total if portfolio_total else 0.0,
            "Explanation": "Shown in Portfolio Composition; excluded from Breakdown and Diversification.",
        },
    ]

    for _, row in composition_df.iloc[:-1].iterrows():
        records.append(
            {
                "Section": "Portfolio Composition",
                "Line Item": row["Asset Classes"],
                "SAA MV": "",
                "TAA MV": "",
                "Total MV": float(row["Portfolio Market Value (CDN)"]),
                "Denominator": portfolio_total,
                "SAA %": "",
                "TAA %": "",
                "Portfolio %": float(row["% of Portfolio (CDN)"]) / 100.0,
                "Explanation": "Total MV / full portfolio market value.",
            }
        )

    for _, row in breakdown_df.iloc[:-1].iterrows():
        saa_pct = float(row["Strategic Asset Allocation %"]) / 100.0
        taa_pct = float(row["Tactical Asset Allocation %"]) / 100.0
        records.append(
            {
                "Section": "Portfolio Breakdown",
                "Line Item": row["Actively Managed Asset Classes"],
                "SAA MV": saa_pct * actively_managed_total,
                "TAA MV": taa_pct * actively_managed_total,
                "Total MV": float(row["Portfolio %"]) / 100.0 * actively_managed_total,
                "Denominator": actively_managed_total,
                "SAA %": saa_pct,
                "TAA %": taa_pct,
                "Portfolio %": float(row["Portfolio %"]) / 100.0,
                "Explanation": "Support-file rows only; separately managed assets excluded.",
            }
        )

    for _, row in diversification_df.iloc[:-1].iterrows():
        saa_pct = float(row["Strategic Asset Allocation %"]) / 100.0
        taa_pct = float(row["Tactical Asset Allocation %"]) / 100.0
        records.append(
            {
                "Section": "Portfolio Diversification",
                "Line Item": row["Portfolio Diversification"],
                "SAA MV": saa_pct * actively_managed_total,
                "TAA MV": taa_pct * actively_managed_total,
                "Total MV": float(row["Portfolio %"]) / 100.0 * actively_managed_total,
                "Denominator": actively_managed_total,
                "SAA %": saa_pct,
                "TAA %": taa_pct,
                "Portfolio %": float(row["Portfolio %"]) / 100.0,
                "Explanation": "Sector-level support-file rows only; separately managed assets excluded.",
            }
        )

    return pd.DataFrame(records)


def build_row_audit(
    section: str,
    rows: pd.DataFrame,
    group_column: str,
    mapping_source_column: str,
    denominator: float,
) -> pd.DataFrame:
    if rows is None or rows.empty:
        return pd.DataFrame(
            columns=[
                "Section",
                "Source Type",
                "Parent Type",
                "Fund Code",
                "Fund Description",
                "Support Code",
                "Support File",
                "Component",
                "Port. Weight %",
                "Parent MV",
                "Weighted MV",
                "SAA MV",
                "TAA MV",
                "Denominator",
                "SAA %",
                "TAA %",
                "Portfolio %",
                "Report Group",
                "Mapping Source",
                "Formula Explanation",
            ]
        )

    detail = rows.copy().reset_index(drop=True)
    records: List[dict] = []
    for idx, row in detail.iterrows():
        excel_row = idx + 5
        parent_type = str(row.get("saa_taa", "")).upper()
        formula_weighted = f"=J{excel_row}*I{excel_row}/100"
        formula_saa_mv = f'=IF(C{excel_row}="SAA",K{excel_row},0)'
        formula_taa_mv = f'=IF(C{excel_row}="TAA",K{excel_row},0)'
        formula_saa_pct = f'=IF(N{excel_row}=0,0,L{excel_row}/N{excel_row})'
        formula_taa_pct = f'=IF(N{excel_row}=0,0,M{excel_row}/N{excel_row})'
        formula_portfolio_pct = f'=IF(N{excel_row}=0,0,K{excel_row}/N{excel_row})'
        records.append(
            {
                "Section": section,
                "Source Type": row.get("Source Type", ""),
                "Parent Type": parent_type,
                "Fund Code": row.get("Fund Code", ""),
                "Fund Description": row.get("Fund Description", ""),
                "Support Code": row.get("support_code", ""),
                "Support File": row.get("Support File", ""),
                "Component": row.get("Component", ""),
                "Port. Weight %": float(row.get("Port. Weight", 0.0)),
                "Parent MV": float(row.get("Total MV (CAD)", 0.0)),
                "Weighted MV": formula_weighted,
                "SAA MV": formula_saa_mv,
                "TAA MV": formula_taa_mv,
                "Denominator": denominator,
                "SAA %": formula_saa_pct,
                "TAA %": formula_taa_pct,
                "Portfolio %": formula_portfolio_pct,
                "Report Group": row.get(group_column, ""),
                "Mapping Source": row.get(mapping_source_column, ""),
                "Formula Explanation": "Weighted MV = Parent MV * Port. Weight / 100; percentages divide by the section denominator.",
            }
        )

    return pd.DataFrame(records)


def add_audit_sheets(
    workbook,
    holdings_df: pd.DataFrame,
    composition_df: pd.DataFrame,
    breakdown_df: pd.DataFrame,
    diversification_df: pd.DataFrame,
    comp_detail: Optional[pd.DataFrame],
    diversification_detail: Optional[pd.DataFrame],
) -> None:
    portfolio_total = float(holdings_df["Total MV (CAD)"].sum())
    actively_managed_total = float(
        composition_df.iloc[-1].get("Actively Managed Market Value (CDN)", portfolio_total)
    )

    summary_ws = workbook.create_sheet("Audit Summary")
    summary_df = build_summary_audit(holdings_df, composition_df, breakdown_df, diversification_df)
    write_audit_sheet(
        summary_ws,
        "Audit Summary",
        summary_df,
        currency_columns=["SAA MV", "TAA MV", "Total MV", "Denominator"],
        percent_columns=["SAA %", "TAA %", "Portfolio %"],
    )

    if comp_detail is None or comp_detail.empty:
        return

    composition_rows = comp_detail[
        comp_detail["Composition Group"].astype(str).str.strip().ne("")
        & select_support_rows_by_hierarchy(comp_detail, "composition")
    ].copy()
    breakdown_rows = comp_detail[
        comp_detail["Source Type"].eq("Support File")
        & comp_detail["Breakdown Group"].astype(str).str.strip().ne("")
        & select_support_rows_by_hierarchy(comp_detail, "breakdown")
    ].copy()

    composition_detail = build_row_audit(
        "Portfolio Composition",
        composition_rows,
        "Composition Group",
        "Composition Mapping Source",
        portfolio_total,
    )
    breakdown_detail = build_row_audit(
        "Portfolio Breakdown",
        breakdown_rows,
        "Breakdown Group",
        "Breakdown Mapping Source",
        actively_managed_total,
    )
    diversification_detail_df = build_row_audit(
        "Portfolio Diversification",
        diversification_detail if diversification_detail is not None else pd.DataFrame(),
        "Diversification Group",
        "Diversification Mapping Source",
        actively_managed_total,
    )

    for sheet_name, title, detail_df in [
        ("Audit Composition", "Audit Detail - Portfolio Composition", composition_detail),
        ("Audit Breakdown", "Audit Detail - Portfolio Breakdown", breakdown_detail),
        ("Audit Diversification", "Audit Detail - Portfolio Diversification", diversification_detail_df),
    ]:
        worksheet = workbook.create_sheet(sheet_name)
        write_audit_sheet(
            worksheet,
            title,
            detail_df,
            currency_columns=["Parent MV", "Weighted MV", "SAA MV", "TAA MV", "Denominator"],
            percent_columns=["SAA %", "TAA %", "Portfolio %"],
        )


def build_excel_report(
    holdings_df: pd.DataFrame,
    composition_df: pd.DataFrame,
    breakdown_df: pd.DataFrame,
    diversification_df: pd.DataFrame,
    composition_fig: go.Figure,
    breakdown_fig: go.Figure,
    diversification_fig: go.Figure,
    reporting_period: Optional[str],
    comp_detail: Optional[pd.DataFrame] = None,
    diversification_detail: Optional[pd.DataFrame] = None,
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font
    except ImportError as exc:
        raise ImportError(
            "Excel export requires `openpyxl`. Install it with `pip install openpyxl`."
        ) from exc

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    holdings_ws = workbook.create_sheet("Holdings")
    holdings_export = holdings_df.copy()
    holding_columns = ["Fund Code", "Fund Description", "Total MV (CAD)", "saa_taa"]
    holdings_export = holdings_export[[col for col in holding_columns if col in holdings_export.columns]].copy()
    holdings_export["Total MV (CAD)"] = holdings_export["Total MV (CAD)"].astype(float)
    write_table(
        holdings_ws,
        start_row=1,
        dataframe=holdings_export,
        currency_columns=["Total MV (CAD)"],
        title="Holdings",
    )

    composition_ws = workbook.create_sheet("Composition")
    composition_export = composition_df.copy()
    composition_export["% of Portfolio (CDN)"] = composition_export["% of Portfolio (CDN)"] / 100.0
    composition_currency_columns = [
        col
        for col in [
            "Portfolio Market Value (CDN)",
            "Actively Managed Market Value (CDN)",
            "Separately Managed Market Value (CDN)",
        ]
        if col in composition_export.columns
    ]
    comp_end_row = write_table(
        composition_ws,
        start_row=1,
        dataframe=composition_export,
        currency_columns=composition_currency_columns,
        percent_columns=["% of Portfolio (CDN)"],
        title="Portfolio Composition",
    )

    breakdown_ws = workbook.create_sheet("Breakdown")
    breakdown_export = breakdown_df.copy()
    for col in ["Strategic Asset Allocation %", "Tactical Asset Allocation %", "Portfolio %"]:
        breakdown_export[col] = breakdown_export[col] / 100.0
    break_end_row = write_table(
        breakdown_ws,
        start_row=1,
        dataframe=breakdown_export,
        percent_columns=["Strategic Asset Allocation %", "Tactical Asset Allocation %", "Portfolio %"],
        title="Portfolio Breakdown",
    )
    next_note_row = break_end_row + 2
    reporting_period_note = build_reporting_period_note(reporting_period)
    if reporting_period_note:
        note_one = breakdown_ws.cell(
            row=next_note_row,
            column=1,
            value=reporting_period_note,
        )
        note_one.font = Font(italic=True)
        next_note_row += 1
    note_two = breakdown_ws.cell(
        row=next_note_row,
        column=1,
        value="The asset class 'Other' may include: Commodities, Derivatives, and/or Preferred Shares.",
    )
    note_two.font = Font(italic=True)
    next_note_row += 1
    note_three = breakdown_ws.cell(
        row=next_note_row,
        column=1,
        value="The above portfolio breakdown does not include separately managed assets which are rebalanced separately from your actively managed assets.",
    )
    note_three.font = Font(italic=True)

    if PILImage is not None:
        comp_png = maybe_render_figure_png(composition_fig, width=1300, height=900)
        break_png = maybe_render_figure_png(breakdown_fig, width=1500, height=900)

        if comp_png:
            image = XLImage(PILImage.open(BytesIO(comp_png)))
            image.width = 820
            image.height = 568
            composition_ws.add_image(image, f"A{comp_end_row + 2}")

        if break_png:
            image = XLImage(PILImage.open(BytesIO(break_png)))
            image.width = 940
            image.height = 564
            breakdown_ws.add_image(image, f"A{break_end_row + 5}")

    diversification_ws = workbook.create_sheet("Diversification")
    diversification_ws.cell(row=1, column=1, value="Portfolio Breakdown (continued)").font = Font(size=14, bold=True)
    diversification_ws.cell(row=3, column=1, value="Portfolio Diversification").font = Font(size=12, bold=True)
    diversification_export = diversification_df.copy()
    for col in ["Strategic Asset Allocation %", "Tactical Asset Allocation %", "Portfolio %"]:
        diversification_export[col] = diversification_export[col] / 100.0
    diversification_end_row = write_table(
        diversification_ws,
        start_row=5,
        dataframe=diversification_export,
        percent_columns=["Strategic Asset Allocation %", "Tactical Asset Allocation %", "Portfolio %"],
    )
    next_div_note_row = diversification_end_row + 2
    if reporting_period_note:
        note_one = diversification_ws.cell(
            row=next_div_note_row,
            column=1,
            value=reporting_period_note,
        )
        note_one.font = Font(italic=True)
        next_div_note_row += 1
    note_two = diversification_ws.cell(
        row=next_div_note_row,
        column=1,
        value="The category 'Other' may include: Commodities, Derivatives, Preferred Shares and/or any sector allocation under 2%.",
    )
    note_two.font = Font(italic=True)
    next_div_note_row += 1
    note_three = diversification_ws.cell(
        row=next_div_note_row,
        column=1,
        value="The above portfolio breakdown does not include separately managed assets which are rebalanced separately from your actively managed assets.",
    )
    note_three.font = Font(italic=True)

    if PILImage is not None:
        diversification_png = maybe_render_figure_png(diversification_fig, width=1800, height=1100)
        if diversification_png:
            image = XLImage(PILImage.open(BytesIO(diversification_png)))
            image.width = 1040
            image.height = 636
            diversification_ws.add_image(image, f"A{diversification_end_row + 5}")

    add_audit_sheets(
        workbook,
        holdings_df,
        composition_df,
        breakdown_df,
        diversification_df,
        comp_detail,
        diversification_detail,
    )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def show_validation_messages(warnings: Iterable[str], info: Iterable[str]) -> None:
    for message in warnings:
        st.warning(message)


CALCULATION_SESSION_KEY = "latest_calculation"


def clear_latest_calculation() -> None:
    st.session_state.pop(CALCULATION_SESSION_KEY, None)


def reset_holding_inputs(clear_files: bool = False, clear_draft: bool = False) -> None:
    if clear_draft:
        clear_draft_state()
    clear_latest_calculation()
    st.session_state["widget_reset_nonce"] = st.session_state.get("widget_reset_nonce", 0) + 1
    st.session_state["holdings_rows"] = pad_holding_rows([])
    st.session_state["holdings_paste_text"] = HOLDINGS_TEXT_TEMPLATE
    if clear_files:
        st.session_state["saved_support_files"] = []
        st.session_state["saved_sma_override_file"] = None
        st.session_state["saved_factset_model_file"] = None
    for key in list(st.session_state.keys()):
        if key.startswith("holding_") or key.startswith("support_files_uploader_") or key.startswith("sma_grouping_uploader_") or key.startswith("factset_model_uploader_") or key.startswith("holdings_paste_area_") or key in {
            "holdings_paste_area",
            "show_audit_view",
            "support_files_uploader",
            "sma_grouping_uploader",
            "factset_model_uploader",
        }:
            del st.session_state[key]


def store_latest_calculation(results: dict, warnings: List[str], excel_bytes: bytes) -> None:
    st.session_state[CALCULATION_SESSION_KEY] = {
        "results": results,
        "warnings": list(warnings),
        "excel_bytes": excel_bytes,
    }


def render_calculation_results(calculation_state: dict) -> None:
    results = calculation_state["results"]
    excel_bytes = calculation_state["excel_bytes"]
    warnings = calculation_state.get("warnings", [])

    composition_fig = build_composition_chart(results["composition"])
    breakdown_fig = build_breakdown_chart(results["breakdown"])
    diversification_fig = build_diversification_chart(results["diversification"])

    st.success("Calculation completed.")
    show_validation_messages(warnings, [])

    factset_models = results.get("factset_models", [])
    if factset_models:
        factset_df = pd.DataFrame(factset_models)
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.subheader("FactSet Model")
        unique_models = sorted(factset_df["FactSet Model"].dropna().unique().tolist())
        if len(unique_models) == 1:
            model_value = unique_models[0]
            base_code, _, suffix = model_value.partition("_")
            model_col1, model_col2, model_col3 = st.columns(3)
            model_col1.metric("FactSet Model", model_value)
            model_col2.metric("Model Allocation", base_code)
            model_col3.metric("Series Suffix", suffix or "N/A")
        else:
            st.dataframe(factset_df, width="stretch", hide_index=True)
        st.markdown("</div>", unsafe_allow_html=True)

    support_file_matches = results["comp_detail"].loc[
        results["comp_detail"]["Source Type"] == "Support File",
        "support_code",
    ].nunique()
    sma_holdings_used = results["comp_detail"].loc[
        results["comp_detail"]["Source Type"] == "SMA",
        "Fund Code",
    ].nunique()

    reporting_period = results.get("reporting_period") or "Not detected"
    summary_col1, summary_col2, summary_col3, summary_col4, summary_col5 = st.columns(5)
    summary_col1.metric("Holdings Rows Used", len(results["holdings"]))
    summary_col2.metric("Support Files Matched", support_file_matches)
    summary_col3.metric("Total Market Value", format_currency(results["portfolio_total"]))
    summary_col4.metric("SMA Holdings Used", sma_holdings_used)
    summary_col5.metric("Asset Allocation Source Period", reporting_period)
    if results.get("reporting_period"):
        st.caption(
            "The asset-allocation source period comes from the uploaded FactSet support files and may be one month earlier than the statement market-value date."
        )

    composition_tab, breakdown_tab, diversification_tab = st.tabs(
        ["Portfolio Composition", "Portfolio Breakdown", "Portfolio Diversification"]
    )

    with composition_tab:
        st.dataframe(
            build_composition_display(results["composition"]),
            width="stretch",
            hide_index=True,
        )
        st.plotly_chart(composition_fig, use_container_width=True)

    with breakdown_tab:
        st.dataframe(
            build_breakdown_display(results["breakdown"]),
            width="stretch",
            hide_index=True,
        )
        st.plotly_chart(breakdown_fig, use_container_width=True)
        reporting_period_note = build_reporting_period_note(results.get("reporting_period"))
        if reporting_period_note:
            st.caption(reporting_period_note)
        st.caption("The asset class 'Other' may include: Commodities, Derivatives, and/or Preferred Shares.")
        st.caption(
            "The above portfolio breakdown does not include separately managed assets which are rebalanced separately from your actively managed assets."
        )

    with diversification_tab:
        st.markdown("#### Portfolio Breakdown (continued)")
        st.subheader("Portfolio Diversification")
        st.write(
            "This section shows the sector-level allocation of the underlying securities split between strategic and tactical asset allocation."
        )
        st.dataframe(
            build_diversification_display(results["diversification"]),
            width="stretch",
            hide_index=True,
        )
        st.plotly_chart(diversification_fig, use_container_width=True)
        reporting_period_note = build_reporting_period_note(results.get("reporting_period"))
        if reporting_period_note:
            st.caption(reporting_period_note)
        st.caption(
            "The category 'Other' may include: Commodities, Derivatives, Preferred Shares and/or any sector allocation under 2%."
        )
        st.caption(
            "The above portfolio breakdown does not include separately managed assets which are rebalanced separately from your actively managed assets."
        )

    st.download_button(
        "Download Full Report as Excel",
        data=excel_bytes,
        file_name="portfolio_composition_report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )

    show_audit = st.checkbox(
        "Show Audit View",
        value=st.session_state.get("show_audit_view", False),
        key="show_audit_view",
        help="Internal review tables for mapping validation. Hidden by default for normal client use.",
    )
    if show_audit:
        audit = build_audit_tables(results)
        st.markdown('<div class="section-card">', unsafe_allow_html=True)
        st.subheader("Audit View")
        audit_col1, audit_col2, audit_col3 = st.columns(3)
        audit_col1.metric("Mapped Labels", audit["summary"]["Mapped Labels"])
        audit_col2.metric("Other Review Rows", audit["summary"]["Other Review Rows"])
        audit_col3.metric("Large Unreviewed Labels", audit["summary"]["Large Unreviewed Labels"])

        st.markdown("**Component Mapping Audit**")
        st.dataframe(
            build_mapping_audit_display(audit["mapping_audit"]),
            width="stretch",
            hide_index=True,
        )

        st.markdown("**Other / Residual Review**")
        st.dataframe(
            build_other_review_display(audit["other_review"]),
            width="stretch",
            hide_index=True,
        )

        st.markdown("**Large Unreviewed Labels**")
        st.dataframe(
            build_unreviewed_labels_display(audit["unreviewed_labels"]),
            width="stretch",
            hide_index=True,
        )
        st.markdown("</div>", unsafe_allow_html=True)


def is_docs_route() -> bool:
    try:
        url = st.context.url
    except Exception:
        return False
    if isinstance(url, bytes):
        url = url.decode("utf-8", errors="ignore")
    if not isinstance(url, str):
        return False
    path = urlparse(url).path.rstrip("/")
    return path == "/docs"


st.markdown(
    """
    <style>
    .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
    }
    .app-subtitle {
        color: #4b5563;
        margin-top: -0.75rem;
        margin-bottom: 1rem;
        font-size: 1rem;
    }
    .section-card {
        padding: 1rem 1.1rem;
        border: 1px solid #e5e7eb;
        border-radius: 16px;
        background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
    }
    div[data-testid="stSidebar"] button[data-testid="stBaseButton-primary"],
    div[data-testid="stSidebar"] .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #2563eb 0%, #0f766e 100%) !important;
        border: 0 !important;
        color: #ffffff !important;
    }
    div[data-testid="stSidebar"] button[data-testid="stBaseButton-primary"]:hover,
    div[data-testid="stSidebar"] .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, #1d4ed8 0%, #0f766e 100%) !important;
        border: 0 !important;
        color: #ffffff !important;
    }
    div[data-testid="stSidebar"] button[data-testid="stBaseButton-primary"]:focus,
    div[data-testid="stSidebar"] .stButton > button[kind="primary"]:focus {
        box-shadow: 0 0 0 0.2rem rgba(37, 99, 235, 0.25) !important;
        color: #ffffff !important;
    }
    div[data-testid="stSidebar"] [data-testid="stCaptionContainer"] {
        color: #9ca3af;
    }
    div[data-testid="stAlert"] {
        border-radius: 14px;
        border: 1px solid rgba(148, 163, 184, 0.24);
    }
    div[data-testid="stAlert"] [data-testid="stMarkdownContainer"] {
        font-size: 0.95rem;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

if is_docs_route():
    st.title("APP Look-Thru Reporting Manual")
    st.markdown(MANUAL_TEXT)
    st.stop()

st.title("APP Look-Thru Reporting")
st.markdown(
    '<div class="app-subtitle">Portfolio Composition and Portfolio Breakdown.</div>',
    unsafe_allow_html=True,
)
st.info("Upload the matching support files before running a calculation.")

if "draft_initialized" not in st.session_state:
    draft_holdings, draft_support_files, draft_sma_override, draft_factset_override = load_draft_state()
    st.session_state["holdings_rows"] = pad_holding_rows(draft_holdings.to_dict("records"))
    st.session_state["holdings_paste_text"] = holdings_df_to_text(draft_holdings)
    st.session_state["saved_support_files"] = draft_support_files
    st.session_state["saved_sma_override_file"] = draft_sma_override
    st.session_state["saved_factset_model_file"] = draft_factset_override
    st.session_state["draft_initialized"] = True
if "widget_reset_nonce" not in st.session_state:
    st.session_state["widget_reset_nonce"] = 0

apply_queued_history_entry_load()

widget_nonce = st.session_state["widget_reset_nonce"]

with st.sidebar:
    st.header("Support Files")
    support_files = st.file_uploader(
        "Support files",
        type=["xlsx", "xls", "xlsm", "csv", "zip"],
        accept_multiple_files=True,
        key=f"support_files_uploader_{widget_nonce}",
        help="Upload the monthly FactSet ZIP bundles or the specific support files for the holdings entered on the page. Extra files are ignored; missing required files will stop the calculation.",
    )

    active_factset_override = st.session_state.get("saved_factset_model_file")
    active_sma_override = st.session_state.get("saved_sma_override_file")

    with st.expander("Advanced"):
        st.caption("Optional reference files. Built-in references are used when these are blank.")
        factset_model_upload = st.file_uploader(
            "FactSet Model Codes File",
            type=["xlsx", "xls", "xlsm", "csv"],
            accept_multiple_files=False,
            key=f"factset_model_uploader_{widget_nonce}",
            help="Optional. Upload the latest Get FactSet Model Codes file to map client TAA fund codes to support-file mandate codes. If omitted, the built-in reference is used.",
        )
        if factset_model_upload is not None:
            uploaded_factset = normalize_uploaded_record(factset_model_upload)
            is_valid, message = validate_factset_model_file(uploaded_factset)
            if is_valid:
                st.session_state["saved_factset_model_file"] = uploaded_factset
                st.success(message)
            else:
                st.session_state["saved_factset_model_file"] = None
                st.warning(message)
        if st.session_state.get("saved_factset_model_file"):
            if st.button("Use Built-In FactSet Codes", width="stretch", key="clear_factset_model_file"):
                st.session_state["saved_factset_model_file"] = None
                st.session_state["widget_reset_nonce"] = st.session_state.get("widget_reset_nonce", 0) + 1
                st.rerun()

        sma_override_upload = st.file_uploader(
            "SMA Grouping File",
            type=["xlsx", "xls", "xlsm", "csv"],
            accept_multiple_files=False,
            key=f"sma_grouping_uploader_{widget_nonce}",
            help="Optional. Only upload this when you need to update the built-in SMA grouping reference.",
        )
        if sma_override_upload is not None:
            uploaded_sma = normalize_uploaded_record(sma_override_upload)
            is_valid, message = validate_sma_grouping_file(uploaded_sma)
            if is_valid:
                st.session_state["saved_sma_override_file"] = uploaded_sma
                st.success(message)
            else:
                st.session_state["saved_sma_override_file"] = None
                st.warning(message)
        if st.session_state.get("saved_sma_override_file"):
            if st.button("Use Built-In SMA Grouping", width="stretch", key="clear_sma_grouping_file"):
                st.session_state["saved_sma_override_file"] = None
                st.session_state["widget_reset_nonce"] = st.session_state.get("widget_reset_nonce", 0) + 1
                st.rerun()

        st.markdown("**Active references**")
        st.caption(
            f"FactSet: `{active_factset_override['filename']}`"
            if active_factset_override
            else "FactSet: built-in"
        )
        st.caption(
            f"SMA grouping: `{active_sma_override['filename']}`"
            if active_sma_override
            else "SMA grouping: built-in"
        )
        if st.button("Reset Draft and Files", width="stretch", key="reset_saved_draft"):
            reset_holding_inputs(clear_files=True, clear_draft=True)
            st.rerun()

    run_calculation = st.button(
        "Run Calculation",
        type="primary",
        width="stretch",
        disabled=False,
    )
    st.text_input(
        "Account label",
        key="account_label",
        placeholder="Optional name for saved history",
        help="Successful manual calculations are saved to history using this label. Leave blank to auto-name from the detected model.",
    )
    with st.expander("Saved Accounts"):
        history_entries = load_account_history()
        if history_entries:
            selected_history_index = st.selectbox(
                "Historical calculations",
                options=list(range(len(history_entries))),
                format_func=lambda index: format_history_entry(history_entries[index]),
                label_visibility="collapsed",
                key="history_entry_select",
            )
            selected_entry = history_entries[selected_history_index]
            selected_history_id = selected_entry.get("id")
            st.caption(format_history_entry_details(selected_entry))
            hist_col1, hist_col2 = st.columns(2)
            if hist_col1.button("Load", width="stretch", key="load_history_entry"):
                if selected_entry:
                    queue_history_entry_load(selected_history_id)
                    st.rerun()
            if hist_col2.button("Delete", width="stretch", key="delete_history_entry"):
                delete_history_entry(selected_history_id)
                st.rerun()
        else:
            st.caption("No saved account history yet.")

st.markdown('<div class="section-card">', unsafe_allow_html=True)
st.subheader("Enter Holdings")
st.write("Paste the holdings export below, then review or adjust the parsed rows if needed.")

if "holdings_rows" not in st.session_state:
    st.session_state["holdings_rows"] = pad_holding_rows([])
if "holdings_paste_text" not in st.session_state:
    st.session_state["holdings_paste_text"] = HOLDINGS_TEXT_TEMPLATE

st.write("Paste either a tab-separated table or the custody export list with fund code, fund name, and market value. Types default intelligently and can be adjusted after import.")
pasted_text = st.text_area(
    "Paste Holdings",
    value=st.session_state["holdings_paste_text"],
    height=180,
    key=f"holdings_paste_area_{widget_nonce}",
    label_visibility="collapsed",
    placeholder="Paste the holdings export here...",
)
paste_col1, _ = st.columns([1.2, 4.8])
if paste_col1.button("Import Pasted Rows", width="stretch"):
    imported_df = parse_holdings_text(pasted_text)
    imported_df = apply_sma_type_detection(
        imported_df,
        sma_override=st.session_state.get("saved_sma_override_file"),
    )
    st.session_state["holdings_rows"] = pad_holding_rows(imported_df.to_dict("records"))
    st.session_state["holdings_paste_text"] = HOLDINGS_TEXT_TEMPLATE
    st.session_state["widget_reset_nonce"] = st.session_state.get("widget_reset_nonce", 0) + 1
    st.rerun()

control_col1, control_col2, control_col3, control_col4, _ = st.columns([1, 1, 1, 1, 2])
if control_col1.button("Add Row", width="stretch"):
    st.session_state["holdings_rows"].append(blank_holding_row())
    st.rerun()
if control_col2.button("Add 5 Rows", width="stretch"):
    st.session_state["holdings_rows"].extend([blank_holding_row() for _ in range(5)])
    st.rerun()
if control_col3.button("Clear Rows", width="stretch"):
    reset_holding_inputs(clear_files=False, clear_draft=False)
    st.rerun()
if control_col4.button("Start Over", width="stretch"):
    reset_holding_inputs(clear_files=True, clear_draft=True)
    st.rerun()

header_cols = st.columns([1.1, 3.6, 1.3, 1.0])
header_cols[0].markdown("**Fund Code**")
header_cols[1].markdown("**Fund Description**")
header_cols[2].markdown("**Total MV (CAD)**")
header_cols[3].markdown("**Type**")

rendered_rows: List[dict] = []
for idx, row in enumerate(st.session_state["holdings_rows"]):
    cols = st.columns([1.1, 3.6, 1.3, 1.0])
    fund_code = cols[0].text_input(
        f"Fund Code {idx + 1}",
        value=row.get("Fund Code", ""),
        label_visibility="collapsed",
        key=f"holding_fund_code_{idx}_{widget_nonce}",
    )
    fund_description = cols[1].text_input(
        f"Fund Description {idx + 1}",
        value=row.get("Fund Description", ""),
        label_visibility="collapsed",
        key=f"holding_fund_description_{idx}_{widget_nonce}",
    )
    total_mv = cols[2].text_input(
        f"Total MV (CAD) {idx + 1}",
        value="" if row.get("Total MV (CAD)", "") is None else str(row.get("Total MV (CAD)", "")),
        label_visibility="collapsed",
        key=f"holding_total_mv_{idx}_{widget_nonce}",
    )
    current_holding_type = normalize_holding_type(row.get("saa_taa", "SAA")) or "SAA"
    saa_taa = cols[3].selectbox(
        f"Type {idx + 1}",
        options=HOLDING_TYPE_OPTIONS,
        index=HOLDING_TYPE_OPTIONS.index(current_holding_type),
        label_visibility="collapsed",
        key=f"holding_saa_taa_{idx}_{widget_nonce}",
    )
    rendered_rows.append(
        {
            "Fund Code": fund_code,
            "Fund Description": fund_description,
            "Total MV (CAD)": total_mv,
            "saa_taa": saa_taa,
        }
    )

st.session_state["holdings_rows"] = rendered_rows
current_holdings_df = strip_blank_holding_rows(holding_rows_to_df(rendered_rows))
st.markdown("</div>", unsafe_allow_html=True)

if support_files:
    st.session_state["saved_support_files"] = [
        {
            "filename": uploaded.name,
            "bytes": uploaded.getvalue(),
        }
        for uploaded in support_files
    ]

save_draft_state(
    current_holdings_df,
    st.session_state.get("saved_support_files", []),
    st.session_state.get("holdings_paste_text", HOLDINGS_TEXT_TEMPLATE),
    st.session_state.get("saved_sma_override_file"),
    st.session_state.get("saved_factset_model_file"),
)

active_support_files = st.session_state.get("saved_support_files", [])
active_factset_override = st.session_state.get("saved_factset_model_file")
active_sma_override = st.session_state.get("saved_sma_override_file")
should_run_calculation = run_calculation or st.session_state.pop("pending_history_recalculate", False)

if active_support_files or active_sma_override or active_factset_override:
    st.markdown('<div class="section-card">', unsafe_allow_html=True)
    st.subheader("Files Selected")
    for saved in active_support_files:
        st.write(f"- `{saved['filename']}`")
    if active_factset_override:
        st.write(f"- FactSet model codes: `{active_factset_override['filename']}`")
    if active_sma_override:
        st.write(f"- SMA grouping: `{active_sma_override['filename']}`")
    st.markdown("</div>", unsafe_allow_html=True)
elif not should_run_calculation:
    if CALCULATION_SESSION_KEY not in st.session_state:
        st.info("Enter holdings, upload the required files, then click `Run Calculation`.")

if should_run_calculation:
    progress = st.progress(0)
    status = st.empty()
    clear_latest_calculation()

    try:
        status.info("Validating entered holdings...")
        holdings_df, holdings_messages = parse_manual_holdings_input(
            current_holdings_df,
            sma_override=active_sma_override,
            factset_override=active_factset_override,
        )
        progress.progress(20)

        status.info("Indexing uploaded support files...")
        support_map, upload_warnings = build_saved_support_map(active_support_files)
        progress.progress(35)

        status.info("Parsing support files and calculating portfolio composition...")
        results, calc_warnings, calc_info = calculate_reports(
            holdings_df,
            support_map,
            sma_override=active_sma_override,
            factset_override=active_factset_override,
        )
        progress.progress(70)

        composition_fig = build_composition_chart(results["composition"])
        breakdown_fig = build_breakdown_chart(results["breakdown"])
        diversification_fig = build_diversification_chart(results["diversification"])
        progress.progress(85)

        status.info("Building Excel download...")
        excel_bytes = build_excel_report(
            results["holdings"],
            results["composition"],
            results["breakdown"],
            results["diversification"],
            composition_fig,
            breakdown_fig,
            diversification_fig,
            results.get("reporting_period"),
            results.get("comp_detail"),
            results.get("diversification_detail"),
        )
        progress.progress(100)

        all_warnings = holdings_messages["warnings"] + upload_warnings + calc_warnings
        store_latest_calculation(results, all_warnings, excel_bytes)
        if run_calculation:
            save_account_history_entry(
                st.session_state.get("account_label", ""),
                holdings_df,
                active_support_files,
                active_sma_override,
                active_factset_override,
                results,
            )
        status.empty()
        progress.empty()

    except Exception as exc:
        clear_latest_calculation()
        progress.empty()
        status.empty()
        st.error(str(exc))

latest_calculation = st.session_state.get(CALCULATION_SESSION_KEY)
if latest_calculation:
    render_calculation_results(latest_calculation)

st.markdown("---")
st.caption("Version 1.0.0")
