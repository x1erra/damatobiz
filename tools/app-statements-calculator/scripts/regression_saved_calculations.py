from __future__ import annotations

import argparse
import json
import math
import re
from pathlib import Path
from typing import Any


APP_UI_MARKER = '\nst.markdown(\n    """\n    <style>'


def load_calculator_namespace(app_path: Path) -> dict[str, Any]:
    source = app_path.read_text(encoding="utf-8")
    if APP_UI_MARKER not in source:
        raise RuntimeError("Could not find UI marker in app.py; refusing to execute full Streamlit app.")
    prefix = source.split(APP_UI_MARKER, 1)[0]
    namespace: dict[str, Any] = {
        "__file__": str(app_path),
        "__name__": "app_regression_calc",
    }
    exec(compile(prefix, str(app_path), "exec"), namespace)
    return namespace


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return round(value, 8)
    if hasattr(value, "item"):
        try:
            return clean_value(value.item())
        except Exception:
            pass
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return value


def frame_records(df: Any) -> list[dict[str, Any]]:
    if df is None or getattr(df, "empty", True):
        return []
    clean = df.copy()
    clean = clean.where(clean.notnull(), None)
    return [
        {str(key): clean_value(value) for key, value in row.items()}
        for row in clean.to_dict("records")
    ]


def focus_detail_records(comp_detail: Any) -> list[dict[str, Any]]:
    if comp_detail is None or getattr(comp_detail, "empty", True):
        return []
    terms = [
        "CI Alternative Credit Opportunities Fund Series IN",
        "CI Alternative Investment Grade Credit Fund I",
        "CI Alternative Investment Grade Credit Fund (ETF C$ Series)",
        "CI Lawrence Park Alternative Investment Grade Credit Fund",
        "Institutional Fiduciary Tr Money Mkt Ptf",
        "MSILF PRIME PORTFOLIO-INST",
        "CI Private Markets Income Fund",
        "Invesco Premier US Gov Money Ptf",
    ]
    pattern = "|".join(re.escape(term) for term in terms)
    mask = comp_detail["Component"].astype(str).str.contains(pattern, case=False, regex=True, na=False)
    keep_cols = [
        "Component",
        "Port. Weight",
        "Weighted MV (CAD)",
        "Composition Group",
        "Composition Mapping Source",
        "Breakdown Group",
        "Breakdown Mapping Source",
        "Diversification Group",
        "Diversification Mapping Source",
        "Block Label",
        "support_code",
        "saa_taa",
    ]
    cols = [col for col in keep_cols if col in comp_detail.columns]
    return frame_records(comp_detail.loc[mask, cols])


def run_entry(namespace: dict[str, Any], entry: dict[str, Any]) -> dict[str, Any]:
    pd = namespace["pd"]
    decode_saved_record = namespace["decode_saved_record"]
    parse_manual_holdings_input = namespace["parse_manual_holdings_input"]
    build_saved_support_map = namespace["build_saved_support_map"]
    calculate_reports = namespace["calculate_reports"]
    manual_cols = namespace["MANUAL_HOLDINGS_COLUMNS"]

    holdings = pd.DataFrame(entry.get("holdings", []) or [])
    for column in manual_cols:
        if column not in holdings.columns:
            holdings[column] = ""
    holdings = holdings[manual_cols].copy()

    support_files = [
        decoded
        for decoded in (decode_saved_record(record) for record in entry.get("support_files", []) or [])
        if decoded
    ]
    sma_override = decode_saved_record(entry.get("sma_override_file"))
    factset_override = decode_saved_record(entry.get("factset_model_file"))

    parsed_holdings, holding_messages = parse_manual_holdings_input(
        holdings,
        sma_override=sma_override,
        factset_override=factset_override,
    )
    support_map, upload_warnings = build_saved_support_map(support_files)
    results, calc_warnings, calc_info = calculate_reports(
        parsed_holdings,
        support_map,
        sma_override=sma_override,
        factset_override=factset_override,
    )

    return {
        "id": entry.get("id"),
        "label": entry.get("label"),
        "created_at": entry.get("created_at"),
        "portfolio_total": clean_value(results.get("portfolio_total")),
        "reporting_period": results.get("reporting_period"),
        "holdings_count": int(len(parsed_holdings)),
        "support_file_count": int(len(support_files)),
        "warnings": list(holding_messages.get("warnings", [])) + list(upload_warnings) + list(calc_warnings),
        "info": list(holding_messages.get("info", [])) + list(calc_info),
        "composition": frame_records(results.get("composition")),
        "breakdown": frame_records(results.get("breakdown")),
        "diversification": frame_records(results.get("diversification")),
        "focus_detail": focus_detail_records(results.get("comp_detail")),
    }


def safe_filename(value: Any, fallback: str) -> str:
    text = str(value or fallback)
    text = re.sub(r"[^A-Za-z0-9_.-]+", "_", text).strip("._")
    return text[:120] or fallback


def calculate_entry(namespace: dict[str, Any], entry: dict[str, Any]) -> tuple[Any, dict[str, Any], list[str], list[str]]:
    pd = namespace["pd"]
    decode_saved_record = namespace["decode_saved_record"]
    parse_manual_holdings_input = namespace["parse_manual_holdings_input"]
    build_saved_support_map = namespace["build_saved_support_map"]
    calculate_reports = namespace["calculate_reports"]
    manual_cols = namespace["MANUAL_HOLDINGS_COLUMNS"]

    holdings = pd.DataFrame(entry.get("holdings", []) or [])
    for column in manual_cols:
        if column not in holdings.columns:
            holdings[column] = ""
    holdings = holdings[manual_cols].copy()

    support_files = [
        decoded
        for decoded in (decode_saved_record(record) for record in entry.get("support_files", []) or [])
        if decoded
    ]
    sma_override = decode_saved_record(entry.get("sma_override_file"))
    factset_override = decode_saved_record(entry.get("factset_model_file"))

    parsed_holdings, holding_messages = parse_manual_holdings_input(
        holdings,
        sma_override=sma_override,
        factset_override=factset_override,
    )
    support_map, upload_warnings = build_saved_support_map(support_files)
    results, calc_warnings, calc_info = calculate_reports(
        parsed_holdings,
        support_map,
        sma_override=sma_override,
        factset_override=factset_override,
    )
    warnings = list(holding_messages.get("warnings", [])) + list(upload_warnings) + list(calc_warnings)
    info = list(holding_messages.get("info", [])) + list(calc_info)
    return parsed_holdings, results, warnings, info


def export_excel(args: argparse.Namespace) -> None:
    app_path = Path(args.app).resolve()
    history_path = Path(args.history).resolve()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    namespace = load_calculator_namespace(app_path)
    payload = json.loads(history_path.read_text(encoding="utf-8"))
    entries = payload.get("entries", [])
    if args.limit:
        entries = entries[: args.limit]

    manifest: list[dict[str, Any]] = []
    for index, entry in enumerate(entries, 1):
        parsed_holdings, results, warnings, info = calculate_entry(namespace, entry)
        build_excel_report = namespace["build_excel_report"]
        build_composition_chart = namespace["build_composition_chart"]
        build_breakdown_chart = namespace["build_breakdown_chart"]
        build_diversification_chart = namespace["build_diversification_chart"]

        excel_bytes = build_excel_report(
            parsed_holdings,
            results["composition"],
            results["breakdown"],
            results["diversification"],
            build_composition_chart(results["composition"]),
            build_breakdown_chart(results["breakdown"]),
            build_diversification_chart(results["diversification"]),
            results.get("reporting_period"),
            results.get("comp_detail"),
            results.get("diversification_detail"),
        )
        filename = f"{index:02d}_{safe_filename(entry.get('label'), 'saved_calculation')}_{safe_filename(entry.get('id'), str(index))}.xlsx"
        output_path = output_dir / filename
        output_path.write_bytes(excel_bytes)
        manifest.append(
            {
                "index": index,
                "id": entry.get("id"),
                "label": entry.get("label"),
                "created_at": entry.get("created_at"),
                "file": output_path.name,
                "warnings": warnings,
                "info": info,
            }
        )
        print(f"ok {index}/{len(entries)} {entry.get('label')} -> {output_path.name}")

    (output_dir / "manifest.json").write_text(json.dumps(manifest, indent=2, sort_keys=True), encoding="utf-8")
    print(f"wrote {output_dir}")


def snapshot(args: argparse.Namespace) -> None:
    app_path = Path(args.app).resolve()
    history_path = Path(args.history).resolve()
    output_path = Path(args.output).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    namespace = load_calculator_namespace(app_path)
    payload = json.loads(history_path.read_text(encoding="utf-8"))
    entries = payload.get("entries", [])
    if args.limit:
        entries = entries[: args.limit]

    runs: list[dict[str, Any]] = []
    failures: list[dict[str, Any]] = []
    for index, entry in enumerate(entries, 1):
        try:
            runs.append(run_entry(namespace, entry))
            print(f"ok {index}/{len(entries)} {entry.get('label')}")
        except Exception as exc:
            failures.append(
                {
                    "id": entry.get("id"),
                    "label": entry.get("label"),
                    "created_at": entry.get("created_at"),
                    "error": str(exc),
                }
            )
            print(f"fail {index}/{len(entries)} {entry.get('label')}: {exc}")

    output = {
        "app": str(app_path),
        "history": str(history_path),
        "run_count": len(runs),
        "failure_count": len(failures),
        "runs": runs,
        "failures": failures,
    }
    output_path.write_text(json.dumps(output, indent=2, sort_keys=True), encoding="utf-8")
    print(f"wrote {output_path}")


def row_key(row: dict[str, Any], columns: list[str]) -> str:
    for column in columns:
        if column in row:
            return str(row[column])
    return ""


def numeric_deltas(before_rows: list[dict[str, Any]], after_rows: list[dict[str, Any]], key_columns: list[str]) -> list[str]:
    before = {row_key(row, key_columns): row for row in before_rows}
    after = {row_key(row, key_columns): row for row in after_rows}
    lines: list[str] = []
    for key in sorted(set(before) | set(after)):
        b = before.get(key, {})
        a = after.get(key, {})
        numeric_columns = sorted(
            {
                col
                for row in (b, a)
                for col, value in row.items()
                if isinstance(value, (int, float)) and not isinstance(value, bool)
            }
        )
        changes = []
        for col in numeric_columns:
            bv = float(b.get(col) or 0.0)
            av = float(a.get(col) or 0.0)
            delta = av - bv
            if abs(delta) > 0.0001:
                changes.append(f"{col}: {bv:,.6f} -> {av:,.6f} ({delta:+,.6f})")
        if changes:
            lines.append(f"- {key}: " + "; ".join(changes))
    return lines


def compare(args: argparse.Namespace) -> None:
    before = json.loads(Path(args.before).read_text(encoding="utf-8"))
    after = json.loads(Path(args.after).read_text(encoding="utf-8"))
    before_runs = {run["id"]: run for run in before.get("runs", [])}
    after_runs = {run["id"]: run for run in after.get("runs", [])}

    lines = [
        "# APP Calculator Regression Diff",
        "",
        f"Before runs: {before.get('run_count')} ok / {before.get('failure_count')} failed",
        f"After runs: {after.get('run_count')} ok / {after.get('failure_count')} failed",
        "",
    ]

    for run_id in sorted(set(before_runs) | set(after_runs), key=lambda value: before_runs.get(value, after_runs.get(value, {})).get("label", "")):
        b = before_runs.get(run_id)
        a = after_runs.get(run_id)
        if not b or not a:
            lines.append(f"## {run_id}")
            lines.append("- Missing from before or after snapshot.")
            lines.append("")
            continue

        section_lines: list[str] = []
        for section, keys in [
            ("composition", ["Asset Classes"]),
            ("breakdown", ["Actively Managed Asset Classes"]),
            ("diversification", ["Portfolio Diversification"]),
        ]:
            deltas = numeric_deltas(b.get(section, []), a.get(section, []), keys)
            if deltas:
                section_lines.append(f"### {section.title()}")
                section_lines.extend(deltas)
                section_lines.append("")

        focus_before = json.dumps(b.get("focus_detail", []), sort_keys=True)
        focus_after = json.dumps(a.get("focus_detail", []), sort_keys=True)
        if focus_before != focus_after:
            section_lines.append("### Focus Detail Changed")
            section_lines.append(f"- Before rows: {len(b.get('focus_detail', []))}")
            section_lines.append(f"- After rows: {len(a.get('focus_detail', []))}")
            section_lines.append("")

        if set(b.get("warnings", [])) != set(a.get("warnings", [])):
            section_lines.append("### Warnings Changed")
            for warning in sorted(set(b.get("warnings", [])) - set(a.get("warnings", []))):
                section_lines.append(f"- Removed: {warning}")
            for warning in sorted(set(a.get("warnings", [])) - set(b.get("warnings", []))):
                section_lines.append(f"- Added: {warning}")
            section_lines.append("")

        if section_lines:
            lines.append(f"## {a.get('label') or b.get('label')}")
            lines.extend(section_lines)

    Path(args.output).write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    print(f"wrote {args.output}")


def clean_cell_value(value: Any) -> Any:
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return round(value, 10)
    return value


def values_equal(before: Any, after: Any, tolerance: float) -> bool:
    before = clean_cell_value(before)
    after = clean_cell_value(after)
    if isinstance(before, (int, float)) and isinstance(after, (int, float)):
        return abs(float(before) - float(after)) <= tolerance
    return before == after


def compare_excel(args: argparse.Namespace) -> None:
    from openpyxl import load_workbook

    before_dir = Path(args.before_dir).resolve()
    after_dir = Path(args.after_dir).resolve()
    output_path = Path(args.output).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    tolerance = float(args.tolerance)

    before_manifest = json.loads((before_dir / "manifest.json").read_text(encoding="utf-8"))
    after_manifest = json.loads((after_dir / "manifest.json").read_text(encoding="utf-8"))
    before_by_id = {item["id"]: item for item in before_manifest}
    after_by_id = {item["id"]: item for item in after_manifest}

    lines = ["# APP Excel Workbook Regression Diff", ""]
    total_diffs = 0
    changed_workbooks = 0
    missing = sorted(set(before_by_id) ^ set(after_by_id))
    if missing:
        lines.append("## Missing Workbooks")
        for run_id in missing:
            lines.append(f"- {run_id}")
        lines.append("")

    for run_id in sorted(set(before_by_id) & set(after_by_id), key=lambda value: before_by_id[value].get("label", "")):
        before_item = before_by_id[run_id]
        after_item = after_by_id[run_id]
        before_wb = load_workbook(before_dir / before_item["file"], data_only=False)
        after_wb = load_workbook(after_dir / after_item["file"], data_only=False)

        workbook_lines: list[str] = []
        if before_wb.sheetnames != after_wb.sheetnames:
            workbook_lines.append(f"- Sheet names changed: {before_wb.sheetnames} -> {after_wb.sheetnames}")

        for sheet_name in sorted(set(before_wb.sheetnames) | set(after_wb.sheetnames)):
            if sheet_name not in before_wb.sheetnames or sheet_name not in after_wb.sheetnames:
                workbook_lines.append(f"### {sheet_name}")
                workbook_lines.append("- Sheet missing before or after.")
                workbook_lines.append("")
                continue
            before_ws = before_wb[sheet_name]
            after_ws = after_wb[sheet_name]
            max_row = max(before_ws.max_row, after_ws.max_row)
            max_col = max(before_ws.max_column, after_ws.max_column)
            sheet_diffs: list[str] = []
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    before_value = before_ws.cell(row=row, column=col).value
                    after_value = after_ws.cell(row=row, column=col).value
                    if not values_equal(before_value, after_value, tolerance):
                        address = before_ws.cell(row=row, column=col).coordinate
                        sheet_diffs.append(f"{address}: {before_value!r} -> {after_value!r}")
            if sheet_diffs:
                total_diffs += len(sheet_diffs)
                workbook_lines.append(f"### {sheet_name} ({len(sheet_diffs)} cell differences)")
                for item in sheet_diffs[: args.max_cells_per_sheet]:
                    workbook_lines.append(f"- {item}")
                if len(sheet_diffs) > args.max_cells_per_sheet:
                    workbook_lines.append(f"- ... {len(sheet_diffs) - args.max_cells_per_sheet} more")
                workbook_lines.append("")

        if workbook_lines:
            changed_workbooks += 1
            lines.append(f"## {after_item.get('label') or before_item.get('label')} [{run_id}]")
            lines.extend(workbook_lines)

    lines.insert(2, f"Compared workbooks: {len(set(before_by_id) & set(after_by_id))}")
    lines.insert(3, f"Changed workbooks: {changed_workbooks}")
    lines.insert(4, f"Cell differences: {total_diffs}")
    lines.insert(5, "")
    output_path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    print(f"wrote {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser()
    subparsers = parser.add_subparsers(dest="command", required=True)

    snap = subparsers.add_parser("snapshot")
    snap.add_argument("--app", required=True)
    snap.add_argument("--history", required=True)
    snap.add_argument("--output", required=True)
    snap.add_argument("--limit", type=int, default=0)
    snap.set_defaults(func=snapshot)

    export = subparsers.add_parser("export-excel")
    export.add_argument("--app", required=True)
    export.add_argument("--history", required=True)
    export.add_argument("--output-dir", required=True)
    export.add_argument("--limit", type=int, default=0)
    export.set_defaults(func=export_excel)

    comp = subparsers.add_parser("compare")
    comp.add_argument("--before", required=True)
    comp.add_argument("--after", required=True)
    comp.add_argument("--output", required=True)
    comp.set_defaults(func=compare)

    comp_xlsx = subparsers.add_parser("compare-excel")
    comp_xlsx.add_argument("--before-dir", required=True)
    comp_xlsx.add_argument("--after-dir", required=True)
    comp_xlsx.add_argument("--output", required=True)
    comp_xlsx.add_argument("--tolerance", type=float, default=0.0000001)
    comp_xlsx.add_argument("--max-cells-per-sheet", type=int, default=80)
    comp_xlsx.set_defaults(func=compare_excel)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
